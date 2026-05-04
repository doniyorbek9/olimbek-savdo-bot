import os
from flask import Flask
from threading import Thread

app = Flask(__name__)

@app.route('/')
def home():
    return "Bot ishlayapti!"

def run():
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)

Thread(target=run).start()

import asyncio
import logging
import random
import psycopg2
from psycopg2.extras import RealDictCursor
import os
import io
from datetime import datetime, date, timedelta
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardRemove, FSInputFile
)
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image, ImageDraw, ImageFont
from admin_panel import app as web_app

# ===================== CONFIG =====================
BOT_TOKEN = os.environ.get("BOT_TOKEN")
ADMIN_IDS = [7948989650]
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
DATABASE_URL = os.environ.get("DATABASE_URL")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# Bloklangan foydalanuvchi middleware
@dp.message.middleware()
async def blocked_user_middleware(handler, message: types.Message, data):
    user = get_user(message.from_user.id)
    if user and user['is_blocked']:
        await message.answer("🚫 Siz bloklangansiz. Botdan foydalanish imkoniyatingiz cheklandi.")
        return
    return await handler(message, data)

# ===================== STATES =====================
class RegState(StatesGroup):
    phone = State()
    name = State()

class OrderState(StatesGroup):
    address = State()
    location = State()
    payment = State()
    check_photo = State()

class AddShopState(StatesGroup):
    tg_id = State()
    name = State()

class AddProductState(StatesGroup):
    name = State()
    price = State()

class EditProductState(StatesGroup):
    choose = State()
    name = State()
    price = State()

class AddCourierState(StatesGroup):
    tg_id = State()
    name = State()
    phone = State()
    shop_id = State()

class ShopAddCourierState(StatesGroup):
    tg_id = State()
    name = State()
    phone = State()

class PhoneOrderState(StatesGroup):
    client_name = State()
    client_phone = State()
    address = State()
    product = State()
    price = State()
    payment = State()

class PromoState(StatesGroup):
    name = State()
    discount_type = State()
    discount_value = State()
    min_sum = State()
    days = State()
    max_uses = State()

class BroadcastState(StatesGroup):
    target = State()
    message = State()

class ChatState(StatesGroup):
    chatting = State()

class ShopEditState(StatesGroup):
    name = State()
    card = State()
    card_name = State()
    work_time = State()

class VacationState(StatesGroup):
    end_date = State()

class DeleteBotState(StatesGroup):
    password = State()
    confirm = State()

class ExcelUploadState(StatesGroup):
    waiting = State()

class ExcelUpdateState(StatesGroup):
    waiting = State()

class AIOrderState(StatesGroup):
    choosing_shop = State()
    typing = State()
    confirming = State()

class AddPhoneState(StatesGroup):
    phone = State()

class SearchState(StatesGroup):
    query = State()
    search_type = State()

class AdminPercentState(StatesGroup):
    percent = State()
    shop_id = State()

class SecondAdminState(StatesGroup):
    tg_id = State()

class EditShopAdminState(StatesGroup):
    field = State()
    value = State()
    shop_id = State()

class EditCourierState(StatesGroup):
    field = State()
    value = State()
    courier_id = State()

class AdminUserChatState(StatesGroup):
    chatting = State()
    user_tg_id = State()

class AdminOrderState(StatesGroup):
    choosing_shop = State()
    phone = State()
    name = State()
    address = State()
    product = State()
    confirm = State()

class ReportErrorState(StatesGroup):
    waiting_text = State()

# ===================== DATABASE =====================
def get_db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    conn.autocommit = False
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()

    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id BIGINT PRIMARY KEY,
        tg_id BIGINT UNIQUE,
        username TEXT,
        full_name TEXT,
        phone TEXT,
        registered_at TEXT,
        is_blocked INTEGER DEFAULT 0
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS shops (
        id BIGINT PRIMARY KEY,
        owner_tg_id BIGINT,
        name TEXT,
        phone TEXT,
        card_number TEXT,
        work_time TEXT,
        is_open INTEGER DEFAULT 1,
        rating REAL DEFAULT 0,
        rating_count INTEGER DEFAULT 0,
        admin_percent REAL DEFAULT 0,
        created_at TEXT,
        vacation_until TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS products (
        id BIGINT PRIMARY KEY,
        shop_id BIGINT,
        name TEXT,
        price REAL,
        is_available INTEGER DEFAULT 1,
        created_at TEXT
    )''')
    try:
        c.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS is_available INTEGER DEFAULT 1")
    except:
        pass

    c.execute('''CREATE TABLE IF NOT EXISTS couriers (
        id BIGINT PRIMARY KEY,
        tg_id BIGINT UNIQUE,
        full_name TEXT,
        phone TEXT,
        shop_id BIGINT,
        is_busy INTEGER DEFAULT 0,
        is_blocked INTEGER DEFAULT 0,
        is_available INTEGER DEFAULT 1,
        queue_order INTEGER DEFAULT 0,
        registered_at TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS orders (
        id BIGSERIAL PRIMARY KEY,
        order_uid TEXT UNIQUE,
        user_tg_id BIGINT,
        shop_id BIGINT,
        courier_tg_id BIGINT,
        products TEXT,
        total_sum REAL,
        address TEXT,
        latitude REAL,
        longitude REAL,
        payment_type TEXT,
        status TEXT DEFAULT 'pending',
        created_at TEXT,
        confirmed_at TEXT,
        delivered_at TEXT,
        check_photo TEXT,
        promo_code TEXT,
        discount REAL DEFAULT 0,
        source TEXT DEFAULT 'normal'
    )''')
    try:
        c.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS source TEXT DEFAULT 'normal'")
    except:
        pass

    c.execute('''CREATE TABLE IF NOT EXISTS promo_codes (
        id BIGINT PRIMARY KEY,
        code TEXT UNIQUE,
        discount_type TEXT,
        discount_value REAL,
        min_sum REAL,
        days INTEGER DEFAULT 0,
        max_uses INTEGER DEFAULT 0,
        used_count INTEGER DEFAULT 0,
        created_at TEXT,
        expires_at TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS chats (
        id BIGSERIAL PRIMARY KEY,
        from_tg_id BIGINT,
        to_tg_id BIGINT,
        message TEXT,
        chat_type TEXT,
        created_at TEXT,
        order_id TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS monthly_reports (
        id BIGSERIAL PRIMARY KEY,
        shop_id BIGINT,
        month TEXT,
        total_income REAL,
        admin_percent REAL,
        admin_share REAL,
        created_at TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS admins (
        id BIGSERIAL PRIMARY KEY,
        tg_id BIGINT UNIQUE,
        added_at TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS carts (
        tg_id BIGINT PRIMARY KEY,
        shop_id BIGINT,
        items TEXT,
        promo TEXT,
        discount REAL DEFAULT 0
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS active_sessions (
        tg_id BIGINT PRIMARY KEY,
        partner BIGINT,
        chat_type TEXT,
        order_id TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS admin_logs (
        id BIGSERIAL PRIMARY KEY,
        admin_tg_id BIGINT,
        action TEXT,
        details TEXT,
        created_at TEXT
    )''')

    conn.commit()
    conn.close()

def gen_id():
    return random.randint(100000, 999999)

def now_str():
    return datetime.now().strftime("%d.%m.%Y %H:%M")

def log_admin_action(admin_tg_id, action, details=""):
    try:
        conn = get_db()
        c = conn.cursor()
        c.execute("INSERT INTO admin_logs (admin_tg_id, action, details, created_at) VALUES (%s,%s,%s,%s)",
                  (admin_tg_id, action, details, now_str()))
        conn.commit()
        conn.close()
    except:
        pass

def is_admin(tg_id):
    if tg_id in ADMIN_IDS:
        return True
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM admins WHERE tg_id=%s", (tg_id,))
    r = c.fetchone()
    conn.close()
    return r is not None

def get_user(tg_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE tg_id=%s", (tg_id,))
    r = c.fetchone()
    conn.close()
    return r

def get_shop_by_owner(tg_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM shops WHERE owner_tg_id=%s", (tg_id,))
    r = c.fetchone()
    conn.close()
    return r

def get_courier_by_tg(tg_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM couriers WHERE tg_id=%s", (tg_id,))
    r = c.fetchone()
    conn.close()
    return r

def user_role(tg_id):
    if is_admin(tg_id):
        return "admin"
    shop = get_shop_by_owner(tg_id)
    if shop:
        return "shop"
    courier = get_courier_by_tg(tg_id)
    if courier:
        return "courier"
    return "user"

# ===================== KEYBOARDS =====================
def main_menu_kb(tg_id):
    role = user_role(tg_id)
    if role == "admin":
        return admin_main_kb()
    elif role == "shop":
        return shop_main_kb()
    elif role == "courier":
        return courier_main_kb()
    else:
        return user_main_kb()

def user_main_kb():
    kb = ReplyKeyboardBuilder()
    kb.button(text="🏪 Do'konlar")
    kb.button(text="🛒 Savat")
    kb.button(text="👤 Profil")
    kb.button(text="💬 Do'kon egasi")
    kb.button(text="💬 Admin")
    kb.button(text="⚠️ Hatolik")
    kb.adjust(2, 1, 2, 1)
    return kb.as_markup(resize_keyboard=True)

def shop_main_kb():
    kb = ReplyKeyboardBuilder()
    kb.button(text="🔓 Do'konni ochish/yopish")
    kb.button(text="✏️ Do'konni tahrirlash")
    kb.button(text="📦 Mahsulotlar")
    kb.button(text="📅 Bugungi hisobot")
    kb.button(text="📊 30 kunlik hisobot")
    kb.button(text="🏖️ Ta'tilda")
    kb.button(text="📞 Telefon buyurtma")
    kb.button(text="🚚 Kuryer qo'shish")
    kb.adjust(2)
    return kb.as_markup(resize_keyboard=True)

def courier_main_kb():
    kb = ReplyKeyboardBuilder()
    kb.button(text="📦 Buyurtmalarim")
    kb.button(text="📅 Kunlik hisobot")
    kb.button(text="📊 30 kunlik hisobot")
    kb.button(text="💬 Do'kon egasi")
    kb.button(text="💬 Admin")
    kb.button(text="📵 Bandman")
    kb.adjust(2)
    return kb.as_markup(resize_keyboard=True)

def admin_main_kb():
    kb = ReplyKeyboardBuilder()
    kb.button(text="🏪 Do'konlar boshqaruv")
    kb.button(text="👥 Mijozlar")
    kb.button(text="🚚 Kuryerlar")
    kb.button(text="📊 Statistika")
    kb.button(text="📣 Xabar yuborish")
    kb.button(text="🔍 Qidirish")
    kb.button(text="📋 Buyurtmalar")
    kb.button(text="⚠️ Muammoli")
    kb.button(text="💰 Moliya")
    kb.button(text="📥 Excel eksport")
    kb.button(text="🎟️ Promo kodlar")
    kb.button(text="👁️ Monitoring")
    kb.button(text="💬 Chatlar")
    kb.button(text="🚫 Bloklangan")
    kb.button(text="📈 Haftalik hisobot")
    kb.button(text="🔐 Admin loglari")
    kb.button(text="📱 Buyurtma berish")
    kb.button(text="🗑️ Bot ma'lumotlari")
    kb.button(text="🏆 Top mijozlar")
    kb.adjust(2)
    return kb.as_markup(resize_keyboard=True)

def back_kb():
    kb = ReplyKeyboardBuilder()
    kb.button(text="⬅️ Orqaga")
    return kb.as_markup(resize_keyboard=True)

def phone_kb():
    kb = ReplyKeyboardBuilder()
    kb.button(text="📱 Telefon raqamni yuborish", request_contact=True)
    return kb.as_markup(resize_keyboard=True, one_time_keyboard=True)

def location_kb():
    kb = ReplyKeyboardBuilder()
    kb.button(text="📍 Lokatsiya yuborish", request_location=True)
    kb.button(text="⏭️ Skip")
    kb.button(text="⬅️ Orqaga")
    kb.adjust(1)
    return kb.as_markup(resize_keyboard=True, one_time_keyboard=True)

# ===================== RECEIPT GENERATOR =====================
def generate_receipt(order_data, user_data, shop_data):
    img = Image.new('RGB', (600, 800), color='white')
    draw = ImageDraw.Draw(img)

    try:
        font_big = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 28)
        font_med = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 20)
        font_small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 16)
    except:
        font_big = ImageFont.load_default()
        font_med = ImageFont.load_default()
        font_small = ImageFont.load_default()

    draw.rectangle([0, 0, 600, 80], fill='#2ECC71')
    draw.text((300, 40), "CHEK", font=font_big, fill='white', anchor='mm')

    y = 100
    draw.text((300, y), "Olimbek SAVDO", font=font_big, fill='#2C3E50', anchor='mm')
    y += 40
    draw.line([20, y, 580, y], fill='#BDC3C7', width=2)
    y += 20

    info = [
        f"Buyurtma ID: {order_data['order_uid']}",
        f"Vaqt: {order_data['created_at']}",
        f"Mijoz: {user_data.get('full_name', '')}",
        f"Tel: {user_data.get('phone', '')}",
        f"Do'kon: {shop_data.get('name', '')}",
        f"Manzil: {order_data['address']}",
        f"To'lov: {order_data['payment_type']}",
    ]

    for line in info:
        draw.text((30, y), line, font=font_med, fill='#2C3E50')
        y += 32

    y += 10
    draw.line([20, y, 580, y], fill='#BDC3C7', width=2)
    y += 15
    draw.text((30, y), "Mahsulotlar:", font=font_med, fill='#2C3E50')
    y += 30

    products = order_data['products'].split('|')
    for p in products:
        draw.text((50, y), f"- {p}", font=font_small, fill='#34495E')
        y += 25

    y += 10
    draw.line([20, y, 580, y], fill='#BDC3C7', width=2)
    y += 15

    if order_data.get('discount', 0) > 0:
        draw.text((30, y), f"Chegirma: -{order_data['discount']:,.0f} so'm", font=font_med, fill='#E74C3C')
        y += 32

    draw.text((30, y), f"Jami: {order_data['total_sum']:,.0f} so'm", font=font_big, fill='#27AE60')
    y += 50

    draw.rectangle([0, y, 600, y+60], fill='#2ECC71')
    draw.text((300, y+30), "Rahmat! Xarid qilganingiz uchun!", font=font_med, fill='white', anchor='mm')

    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf


def generate_pdf_receipt(order_data, user_data, shop_data):
    try:
        from reportlab.lib.pagesizes import A6
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A6,
                                rightMargin=1*cm, leftMargin=1*cm,
                                topMargin=1*cm, bottomMargin=1*cm)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle('title', parent=styles['Normal'],
                                     fontSize=14, fontName='Helvetica-Bold',
                                     alignment=1, textColor=colors.HexColor('#27AE60'))
        normal_style = ParagraphStyle('norm', parent=styles['Normal'],
                                      fontSize=9, fontName='Helvetica')
        bold_style = ParagraphStyle('bold', parent=styles['Normal'],
                                    fontSize=10, fontName='Helvetica-Bold')

        story.append(Paragraph("OLIMBEK SAVDO", title_style))
        story.append(Paragraph("CHEK / KVITANSIYA", title_style))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
        story.append(Spacer(1, 0.2*cm))

        info = [
            f"Buyurtma ID: #{order_data.get('order_uid', '')}",
            f"Vaqt: {order_data.get('created_at', '')}",
            f"Mijoz: {user_data.get('full_name', '')}",
            f"Telefon: {user_data.get('phone', '')}",
            f"Do'kon: {shop_data.get('name', '')}",
            f"Manzil: {order_data.get('address', '')}",
            f"To'lov: {order_data.get('payment_type', '')}",
        ]
        for line in info:
            story.append(Paragraph(line, normal_style))
            story.append(Spacer(1, 0.1*cm))

        story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
        story.append(Paragraph("Mahsulotlar:", bold_style))
        products = order_data.get('products', '').split('|')
        for p in products:
            story.append(Paragraph(f"- {p.strip()}", normal_style))

        story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
        if order_data.get('discount', 0) > 0:
            story.append(Paragraph(f"Chegirma: -{order_data['discount']:,.0f} so'm", normal_style))
        story.append(Paragraph(f"JAMI: {order_data.get('total_sum', 0):,.0f} so'm", title_style))
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph("Xarid qilganingiz uchun rahmat!", title_style))

        doc.build(story)
        buf.seek(0)
        return buf, "pdf"
    except ImportError:
        return generate_receipt(order_data, user_data, shop_data), "png"


# ===================== HANDLERS =====================

@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    await state.clear()
    tg_id = message.from_user.id
    user = get_user(tg_id)

    if user and user['is_blocked']:
        await message.answer("🚫 Siz bloklangansiz. Botdan foydalanish imkoniyatingiz cheklandi.")
        return

    if user:
        await message.answer(
            f"Xush kelibsiz, {user['full_name']}! 👋",
            reply_markup=main_menu_kb(tg_id)
        )
        return

    role = user_role(tg_id)
    if role in ["shop", "courier", "admin"]:
        await message.answer("Xush kelibsiz! 👋", reply_markup=main_menu_kb(tg_id))
        return

    await message.answer(
        "👋 Salom! Olimbek SAVDO botiga xush kelibsiz!\n\n📱 Telefon raqamingizni yuboring:",
        reply_markup=phone_kb()
    )
    await state.set_state(RegState.phone)

@dp.message(RegState.phone, F.contact)
async def reg_phone(message: types.Message, state: FSMContext):
    await state.update_data(phone=message.contact.phone_number)
    await message.answer("✍️ Ism familyangizni kiriting:", reply_markup=ReplyKeyboardRemove())
    await state.set_state(RegState.name)

@dp.message(RegState.name)
async def reg_name(message: types.Message, state: FSMContext):
    data = await state.get_data()
    tg_id = message.from_user.id
    uid = gen_id()

    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO users (id, tg_id, username, full_name, phone, registered_at) VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (tg_id) DO NOTHING",
                  (uid, tg_id, message.from_user.username or "", message.text, data['phone'], now_str()))
        conn.commit()
    except:
        conn.rollback()
    finally:
        conn.close()

    await state.clear()
    await message.answer(
        f"✅ Ro'yxatdan o'tdingiz!\n👤 {message.text}\n📱 {data['phone']}\n🆔 {uid}",
        reply_markup=user_main_kb()
    )

# --- PROFILE ---
@dp.message(F.text == "👤 Profil")
async def profile(message: types.Message):
    user = get_user(message.from_user.id)
    if not user:
        await message.answer("Avval ro'yxatdan o'ting: /start")
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) as cnt FROM orders WHERE user_tg_id=%s", (message.from_user.id,))
    order_count = c.fetchone()['cnt']
    conn.close()

    text = (f"👤 Profil\n\n"
            f"📛 Ism: {user['full_name']}\n"
            f"📱 Tel: {user['phone']}\n"
            f"🆔 ID: {user['id']}\n"
            f"👤 Username: @{user['username'] or 'yoq'}\n"
            f"📅 Ro'yxat: {user['registered_at']}\n"
            f"📦 Jami buyurtmalar: {order_count}")

    kb = InlineKeyboardBuilder()
    kb.button(text="📦 Buyurtmalar tarixi", callback_data="my_orders")
    await message.answer(text, reply_markup=kb.as_markup())

@dp.callback_query(F.data == "my_orders")
async def my_orders(callback: types.CallbackQuery):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM orders WHERE user_tg_id=%s ORDER BY created_at DESC LIMIT 20", (callback.from_user.id,))
    orders = c.fetchall()
    conn.close()

    if not orders:
        await callback.message.answer("📦 Buyurtmalar yo'q")
        return

    kb = InlineKeyboardBuilder()
    for o in orders:
        status_emoji = {"pending": "⏳", "confirmed": "✅", "on_way": "🚗", "delivered": "✅", "rejected": "❌"}.get(o['status'], "❓")
        kb.button(text=f"{status_emoji} #{o['order_uid']} — {o['created_at']}", callback_data=f"order_detail_{o['order_uid']}")
    kb.adjust(1)
    await callback.message.answer("📦 Buyurtmalar tarixi:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("order_detail_"))
async def order_detail(callback: types.CallbackQuery):
    uid = callback.data.split("_", 2)[2]
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM orders WHERE order_uid=%s", (uid,))
    o = c.fetchone()
    conn.close()

    if not o:
        await callback.message.answer("Buyurtma topilmadi")
        return

    status_map = {"pending": "⏳ Kutilmoqda", "confirmed": "✅ Tasdiqlandi", "on_way": "🚗 Yo'lda",
                  "delivered": "✅ Yetkazildi", "rejected": "❌ Rad etildi"}

    text = (f"📦 Buyurtma #{o['order_uid']}\n\n"
            f"📅 Vaqt: {o['created_at']}\n"
            f"📍 Manzil: {o['address']}\n"
            f"🛍️ Mahsulotlar: {o['products']}\n"
            f"💰 Summa: {o['total_sum']:,.0f} so'm\n"
            f"💳 To'lov: {o['payment_type']}\n"
            f"📊 Holat: {status_map.get(o['status'], o['status'])}\n"
            f"✅ Tasdiqlangan: {o['confirmed_at'] or 'Yo\'q'}\n"
            f"🚚 Yetkazilgan: {o['delivered_at'] or 'Yo\'q'}")

    kb = InlineKeyboardBuilder()
    conn2 = get_db()
    c2 = conn2.cursor()
    c2.execute("SELECT owner_tg_id FROM shops WHERE id=%s", (o['shop_id'],))
    shop = c2.fetchone()
    conn2.close()

    if shop:
        kb.button(text="💬 Do'kon egasi bilan yozish", callback_data=f"chat_shop_{o['shop_id']}_{o['order_uid']}")
    kb.adjust(1)
    await callback.message.answer(text, reply_markup=kb.as_markup())

# --- SHOPS LIST ---
@dp.message(F.text == "🏪 Do'konlar")
async def shops_list(message: types.Message):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM shops WHERE is_open=1")
    shops = c.fetchall()
    conn.close()

    if not shops:
        await message.answer("🏪 Hozircha do'konlar yo'q")
        return

    kb = InlineKeyboardBuilder()
    for s in shops:
        rating_text = f"⭐{s['rating']:.1f}" if s['rating_count'] > 0 else "⭐Yangi"
        kb.button(text=f"🏪 {s['name']} {rating_text}", callback_data=f"shop_{s['id']}")
    kb.adjust(1)
    await message.answer("🏪 Do'konlar ro'yxati:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("shop_") & ~F.data.startswith("shop_edit") & ~F.data.startswith("shop_del"))
async def shop_detail(callback: types.CallbackQuery):
    shop_id = int(callback.data.split("_")[1])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM shops WHERE id=%s", (shop_id,))
    shop = c.fetchone()
    c.execute("SELECT * FROM products WHERE shop_id=%s AND (is_available IS NULL OR is_available=1)", (shop_id,))
    products = c.fetchall()
    conn.close()

    if not shop:
        await callback.message.answer("Do'kon topilmadi")
        return

    if shop['vacation_until']:
        try:
            vac_date = datetime.strptime(shop['vacation_until'], "%d.%m.%Y")
            if vac_date.date() >= date.today():
                await callback.message.answer(f"🏖️ Do'kon ta'tilda — {shop['vacation_until']} gacha")
                return
        except:
            pass

    text = f"🏪 {shop['name']}\n⭐ Reyting: {shop['rating']:.1f} ({shop['rating_count']} ovoz)\n⏰ Ish vaqti: {shop['work_time'] or 'Ko\'rsatilmagan'}\n\n📦 Mahsulotlar:"

    kb = InlineKeyboardBuilder()
    cart = cart_get(callback.from_user.id)
    cart_items = cart.get("items", {}) if cart.get("shop_id") == shop_id else {}
    for p in products:
        qty = cart_items.get(str(p['id']), {}).get('qty', 0)
        qty_text = f" ✅ {qty}x" if qty > 0 else ""
        kb.button(text=f"{p['name']} — {p['price']:,.0f} so'm{qty_text}", callback_data=f"add_cart_{p['id']}_{shop_id}")
    kb.button(text="🛒 Savat", callback_data="view_cart")
    kb.button(text="📞 Telefon buyurtma", callback_data=f"phone_order_{shop_id}")
    kb.button(text="⬅️ Orqaga", callback_data="back_shops")
    kb.adjust(1)
    await callback.message.answer(text, reply_markup=kb.as_markup())

@dp.callback_query(F.data == "back_shops")
async def back_shops(callback: types.CallbackQuery):
    await callback.message.delete()
    await shops_list(callback.message)

# --- CART ---
def cart_get(tg_id):
    import json
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM carts WHERE tg_id=%s", (tg_id,))
    row = c.fetchone()
    conn.close()
    if not row:
        return {}
    return {
        "shop_id": row["shop_id"],
        "items": json.loads(row["items"] or "{}"),
        "promo": row["promo"] or "",
        "discount": row["discount"] or 0
    }

def cart_save(tg_id, data):
    import json
    conn = get_db()
    c = conn.cursor()
    c.execute('''INSERT INTO carts (tg_id, shop_id, items, promo, discount)
                 VALUES (%s,%s,%s,%s,%s)
                 ON CONFLICT (tg_id) DO UPDATE SET
                 shop_id=EXCLUDED.shop_id, items=EXCLUDED.items,
                 promo=EXCLUDED.promo, discount=EXCLUDED.discount''',
              (tg_id, data.get("shop_id"), json.dumps(data.get("items", {})),
               data.get("promo", ""), data.get("discount", 0)))
    conn.commit()
    conn.close()

def cart_clear(tg_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM carts WHERE tg_id=%s", (tg_id,))
    conn.commit()
    conn.close()

@dp.callback_query(F.data.startswith("add_cart_"))
async def add_to_cart(callback: types.CallbackQuery):
    parts = callback.data.split("_")
    prod_id = int(parts[2])
    shop_id = int(parts[3])
    tg_id = callback.from_user.id

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM products WHERE id=%s", (prod_id,))
    product = c.fetchone()
    conn.close()

    if not product:
        await callback.answer("Mahsulot topilmadi")
        return

    cart = cart_get(tg_id)
    if not cart or cart.get("shop_id") != shop_id:
        cart = {"shop_id": shop_id, "items": {}, "promo": "", "discount": 0}

    items = cart["items"]
    key = str(prod_id)
    if key in items:
        items[key]["qty"] += 1
    else:
        items[key] = {"name": product['name'], "price": product['price'], "qty": 1}
    cart["items"] = items
    cart_save(tg_id, cart)

    total_qty = sum(i['qty'] for i in cart["items"].values())
    await callback.answer(f"✅ {product['name']} qo'shildi! Savatda: {total_qty} ta")

    conn2 = get_db()
    c2 = conn2.cursor()
    c2.execute("SELECT * FROM shops WHERE id=%s", (shop_id,))
    shop = c2.fetchone()
    c2.execute("SELECT * FROM products WHERE shop_id=%s AND (is_available IS NULL OR is_available=1)", (shop_id,))
    products_list = c2.fetchall()
    conn2.close()

    if shop:
        text = f"🏪 {shop['name']}\n⭐ Reyting: {shop['rating']:.1f} ({shop['rating_count']} ovoz)\n⏰ Ish vaqti: {shop['work_time'] or 'Ko\'rsatilmagan'}\n\n📦 Mahsulotlar:"
        kb2 = InlineKeyboardBuilder()
        cart_items = cart.get("items", {})
        for p in products_list:
            qty = cart_items.get(str(p['id']), {}).get('qty', 0)
            qty_text = f" 🛒 {qty}x" if qty > 0 else ""
            kb2.button(text=f"{p['name']} — {p['price']:,.0f} so'm{qty_text}", callback_data=f"add_cart_{p['id']}_{shop_id}")
        kb2.button(text="🛒 Savat", callback_data="view_cart")
        kb2.button(text="📞 Telefon buyurtma", callback_data=f"phone_order_{shop_id}")
        kb2.button(text="⬅️ Orqaga", callback_data="back_shops")
        kb2.adjust(1)
        try:
            await callback.message.edit_reply_markup(reply_markup=kb2.as_markup())
        except:
            pass

@dp.message(F.text == "🛒 Savat")
async def view_cart_msg(message: types.Message):
    await show_cart(message, message.from_user.id)

@dp.callback_query(F.data == "view_cart")
async def view_cart_cb(callback: types.CallbackQuery):
    await show_cart(callback.message, callback.from_user.id)

async def show_cart(message, tg_id):
    cart_data = cart_get(tg_id)
    items = cart_data.get("items", {})

    if not items:
        await message.answer("🛒 Savat bo'sh")
        return

    text = "🛒 <b>Savat:</b>\n\n"
    total = 0
    kb = InlineKeyboardBuilder()

    for pid, item in items.items():
        subtotal = item['price'] * item['qty']
        total += subtotal
        text += f"▪️ {item['name']}\n   {item['qty']} x {item['price']:,.0f} = <b>{subtotal:,.0f} so'm</b>\n\n"
        kb.button(text=f"➖", callback_data=f"cart_minus_{pid}")
        kb.button(text=f"{item['qty']}x — {item['name']}", callback_data=f"cart_info_{pid}")
        kb.button(text=f"➕", callback_data=f"cart_plus_{pid}")

    discount = cart_data.get("discount", 0)
    promo = cart_data.get("promo", "")
    if promo and discount:
        text += f"🎟️ Promo ({promo}): <b>-{discount:,.0f} so'm</b>\n"
        text += f"💰 Jami: <b>{total - discount:,.0f} so'm</b>"
    else:
        text += f"💰 Jami: <b>{total:,.0f} so'm</b>"

    kb.button(text="🎟️ Promo kod", callback_data="enter_promo")
    kb.button(text="🗑️ Tozalash", callback_data="clear_cart")
    kb.button(text="✅ Buyurtma berish", callback_data="checkout")
    kb.button(text="⬅️ Orqaga", callback_data="back_shops")

    adjusts = [3] * len(items) + [2, 1, 1]
    kb.adjust(*adjusts)

    await message.answer(text, reply_markup=kb.as_markup(), parse_mode="HTML")

async def edit_cart(message, tg_id):
    cart_data = cart_get(tg_id)
    items = cart_data.get("items", {})

    if not items:
        try:
            await message.edit_text("🛒 Savat bo'sh")
        except:
            await message.answer("🛒 Savat bo'sh")
        return

    text = "🛒 <b>Savat:</b>\n\n"
    total = 0
    kb = InlineKeyboardBuilder()

    for pid, item in items.items():
        subtotal = item['price'] * item['qty']
        total += subtotal
        text += f"▪️ {item['name']}\n   {item['qty']} x {item['price']:,.0f} = <b>{subtotal:,.0f} so'm</b>\n\n"
        kb.button(text="➖", callback_data=f"cart_minus_{pid}")
        kb.button(text=f"{item['qty']}x — {item['name']}", callback_data=f"cart_info_{pid}")
        kb.button(text="➕", callback_data=f"cart_plus_{pid}")

    discount = cart_data.get("discount", 0)
    promo = cart_data.get("promo", "")
    if promo and discount:
        text += f"🎟️ Promo ({promo}): <b>-{discount:,.0f} so'm</b>\n"
        text += f"💰 Jami: <b>{total - discount:,.0f} so'm</b>"
    else:
        text += f"💰 Jami: <b>{total:,.0f} so'm</b>"

    kb.button(text="🎟️ Promo kod", callback_data="enter_promo")
    kb.button(text="🗑️ Tozalash", callback_data="clear_cart")
    kb.button(text="✅ Buyurtma berish", callback_data="checkout")
    kb.button(text="⬅️ Orqaga", callback_data="back_shops")

    adjusts = [3] * len(items) + [2, 1, 1]
    kb.adjust(*adjusts)

    try:
        await message.edit_text(text, reply_markup=kb.as_markup(), parse_mode="HTML")
    except:
        await message.answer(text, reply_markup=kb.as_markup(), parse_mode="HTML")

@dp.callback_query(F.data.startswith("cart_info_"))
async def cart_info(callback: types.CallbackQuery):
    await callback.answer()

@dp.callback_query(F.data.startswith("cart_plus_"))
async def cart_plus(callback: types.CallbackQuery):
    pid = str(callback.data.split("_")[2])
    tg_id = callback.from_user.id
    cart = cart_get(tg_id)
    if cart and pid in cart.get("items", {}):
        cart["items"][pid]["qty"] += 1
        cart_save(tg_id, cart)
    await callback.answer("➕ Qo'shildi")
    await edit_cart(callback.message, tg_id)

@dp.callback_query(F.data.startswith("cart_minus_"))
async def cart_minus(callback: types.CallbackQuery):
    pid = str(callback.data.split("_")[2])
    tg_id = callback.from_user.id
    cart = cart_get(tg_id)
    if cart and pid in cart.get("items", {}):
        if cart["items"][pid]["qty"] > 1:
            cart["items"][pid]["qty"] -= 1
        else:
            del cart["items"][pid]
        cart_save(tg_id, cart)
    await callback.answer("➖ Kamaytirildi")
    await edit_cart(callback.message, tg_id)

@dp.callback_query(F.data == "clear_cart")
async def clear_cart(callback: types.CallbackQuery):
    cart_clear(callback.from_user.id)
    await callback.message.answer("🗑️ Savat tozalandi")

@dp.callback_query(F.data == "enter_promo")
async def enter_promo(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.answer("🎟️ Promo kodni kiriting:")
    await state.set_state(PromoState.name)

@dp.message(PromoState.name)
async def check_promo(message: types.Message, state: FSMContext):
    code = message.text.strip().upper()
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM promo_codes WHERE code=%s", (code,))
    promo = c.fetchone()
    conn.close()

    tg_id = message.from_user.id
    cart_data = cart_get(tg_id)
    items = cart_data.get("items", {})
    total = sum(i['price'] * i['qty'] for i in items.values())

    if not promo:
        await message.answer("❌ Promo kod topilmadi")
        await state.clear()
        return

    if promo['max_uses'] > 0 and promo['used_count'] >= promo['max_uses']:
        await message.answer("❌ Promo kod tugagan")
        await state.clear()
        return

    if promo['expires_at']:
        try:
            exp = datetime.strptime(promo['expires_at'], "%d.%m.%Y")
            if exp.date() < date.today():
                await message.answer("❌ Promo kod muddati tugagan")
                await state.clear()
                return
        except:
            pass

    if total < promo['min_sum']:
        await message.answer(f"❌ Minimal summa: {promo['min_sum']:,.0f} so'm")
        await state.clear()
        return

    if promo['discount_type'] == 'percent':
        discount = total * promo['discount_value'] / 100
    else:
        discount = promo['discount_value']

    cart_data['promo'] = code
    cart_data['discount'] = discount
    cart_save(tg_id, cart_data)

    await message.answer(f"✅ Promo kod qo'llandi! Chegirma: -{discount:,.0f} so'm")
    await state.clear()

# --- CHECKOUT ---
@dp.callback_query(F.data == "checkout")
async def checkout(callback: types.CallbackQuery, state: FSMContext):
    tg_id = callback.from_user.id
    user = get_user(tg_id)
    if not user:
        await callback.message.answer("Avval ro'yxatdan o'ting: /start")
        return

    cart_data = cart_get(tg_id)
    if not cart_data.get("items"):
        await callback.message.answer("🛒 Savat bo'sh")
        return

    await callback.message.answer("📍 Manzilingizni kiriting (matn):", reply_markup=back_kb())
    await state.set_state(OrderState.address)

@dp.message(OrderState.address)
async def order_address(message: types.Message, state: FSMContext):
    if message.text == "⬅️ Orqaga":
        await state.clear()
        await message.answer("Bosh menyu:", reply_markup=user_main_kb())
        return
    await state.update_data(address=message.text)
    await message.answer("📍 Lokatsiyangizni yuboring:", reply_markup=location_kb())
    await state.set_state(OrderState.location)

@dp.message(OrderState.location, F.location)
async def order_location(message: types.Message, state: FSMContext):
    await state.update_data(lat=message.location.latitude, lon=message.location.longitude)
    kb = InlineKeyboardBuilder()
    kb.button(text="💵 Naqd", callback_data="pay_cash")
    kb.button(text="💳 Karta", callback_data="pay_card")
    await message.answer("💳 To'lov turini tanlang:", reply_markup=kb.as_markup())
    await state.set_state(OrderState.payment)

@dp.message(OrderState.location)
async def order_location_skip(message: types.Message, state: FSMContext):
    if message.text == "⬅️ Orqaga":
        await state.clear()
        await message.answer("Bosh menyu:", reply_markup=user_main_kb())
        return
    await state.update_data(lat=None, lon=None)
    kb = InlineKeyboardBuilder()
    kb.button(text="💵 Naqd", callback_data="pay_cash")
    kb.button(text="💳 Karta", callback_data="pay_card")
    await message.answer("💳 To'lov turini tanlang:", reply_markup=kb.as_markup())
    await state.set_state(OrderState.payment)

@dp.callback_query(F.data.in_(["pay_cash", "pay_card"]), StateFilter(OrderState.payment))
async def order_payment(callback: types.CallbackQuery, state: FSMContext):
    await state.update_data(payment=callback.data)
    tg_id = callback.from_user.id

    if callback.data == "pay_card":
        cart_data = cart_get(tg_id)
        shop_id = cart_data.get("shop_id")
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT card_number FROM shops WHERE id=%s", (shop_id,))
        shop = c.fetchone()
        conn.close()

        card_raw = shop['card_number'] if shop and shop['card_number'] else "Karta raqami kiritilmagan"
        if " | " in card_raw:
            card_num, card_owner = card_raw.split(" | ", 1)
            card_display = card_num + "  |  " + card_owner
        else:
            card_display = card_raw
        await callback.message.answer(
            "💳 Karta raqami: <code>" + card_display + "</code>\n\nShu kartaga pul o'tkazing va chek rasmini yuboring:",
            parse_mode="HTML"
        )
        await state.set_state(OrderState.check_photo)
    else:
        await finalize_order(callback.message, state, tg_id, "Naqd", None)

@dp.message(OrderState.check_photo, F.photo)
async def order_check_photo(message: types.Message, state: FSMContext):
    photo_id = message.photo[-1].file_id
    await finalize_order(message, state, message.from_user.id, "Karta", photo_id)

@dp.message(OrderState.check_photo, F.document)
async def order_check_photo_doc(message: types.Message, state: FSMContext):
    photo_id = message.document.file_id
    await finalize_order(message, state, message.from_user.id, "Karta", photo_id)

@dp.message(OrderState.check_photo)
async def order_check_photo_wrong(message: types.Message, state: FSMContext):
    await message.answer("📸 Iltimos, chek rasmini yuboring (rasm ko'rinishida)")

async def finalize_order(message, state, tg_id, payment_type, check_photo_id, source='normal'):
    try:
        data = await state.get_data()
        source = data.get('order_source', source)
        cart_data = cart_get(tg_id)
        items = cart_data.get("items", {})
        shop_id = cart_data.get("shop_id")
        discount = cart_data.get("discount", 0)
        promo = cart_data.get("promo", "")

        if not items or not shop_id:
            await message.answer("❌ Savat bo'sh yoki do'kon tanlanmagan")
            await state.clear()
            return

        total = sum(i['price'] * i['qty'] for i in items.values()) - discount
        products_str = "|".join([f"{i['name']} x{i['qty']} ({i['price']:,.0f})" for i in items.values()])
        products_display = "\n".join([f"  • {i['name']} x{i['qty']} — {i['price']:,.0f} so'm" for i in items.values()])

        order_uid = str(gen_id())
        conn = get_db()
        c = conn.cursor()
        order_id = gen_id()
        c.execute('''INSERT INTO orders (id, order_uid, user_tg_id, shop_id, products, total_sum, address,
                  latitude, longitude, payment_type, status, created_at, check_photo, promo_code, discount, source)
                  VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)''',
                  (order_id, order_uid, tg_id, shop_id, products_str, total,
                   data.get('address', ''), data.get('lat'), data.get('lon'),
                   payment_type, 'pending', now_str(), check_photo_id, promo, discount, source))

        if promo:
            c.execute("UPDATE promo_codes SET used_count=used_count+1 WHERE code=%s", (promo,))

        c.execute("SELECT * FROM shops WHERE id=%s", (shop_id,))
        shop = c.fetchone()
        conn.commit()
        conn.close()

        user = get_user(tg_id)

        try:
            order_row = type('obj', (object,), {
                'order_uid': order_uid, 'created_at': now_str(),
                'address': data.get('address', ''), 'products': products_str,
                'total_sum': total, 'payment_type': payment_type,
                'discount': discount
            })()
            receipt_buf = generate_receipt(order_row, dict(user) if user else {}, dict(shop) if shop else {})
            await message.answer_photo(
                types.BufferedInputFile(receipt_buf.read(), filename="chek.png"),
                caption=f"✅ Buyurtmangiz qabul qilindi!\n📦 ID: #{order_uid}",
                reply_markup=user_main_kb()
            )
        except Exception as e:
            logger.error(f"Receipt xatosi: {e}")
            await message.answer(
                f"✅ Buyurtmangiz qabul qilindi!\n📦 ID: #{order_uid}",
                reply_markup=user_main_kb()
            )

        if shop:
            kb = InlineKeyboardBuilder()
            kb.button(text="✅ Tasdiqlash", callback_data=f"confirm_order_{order_uid}")
            kb.button(text="❌ Rad etish", callback_data=f"reject_order_{order_uid}")
            kb.adjust(2)

            source_badge = "🤖 AI orqali buyurtma\n\n" if source == "ai" else ""
            admin_text = (source_badge + f"🆕 Yangi buyurtma #{order_uid}\n\n"
                          f"👤 Mijoz: {user['full_name'] if user else 'Noma\'lum'}\n"
                          f"📱 Tel: {user['phone'] if user else ''}\n"
                          f"🆔 TG ID: {tg_id}\n"
                          f"👤 Username: @{user['username'] if user and user['username'] else 'yoq'}\n"
                          f"📍 Manzil: {data.get('address', '')}\n"
                          f"🛍️ Mahsulotlar:\n{products_display}\n"
                          f"💰 Jami: {total:,.0f} so'm\n"
                          f"💳 To'lov: {payment_type}\n"
                          f"📅 Vaqt: {now_str()}")

            try:
                if check_photo_id:
                    await bot.send_photo(shop['owner_tg_id'], check_photo_id, caption=admin_text, reply_markup=kb.as_markup())
                else:
                    await bot.send_message(shop['owner_tg_id'], admin_text, reply_markup=kb.as_markup())
                if data.get('lat') and data.get('lon'):
                    await bot.send_location(shop['owner_tg_id'], data['lat'], data['lon'])
            except Exception as e:
                logger.error(f"Do'kon egasiga xabar yuborilmadi: {e}")
                for admin_id in ADMIN_IDS:
                    try:
                        await bot.send_message(admin_id, "⚠️ Do'kon egasiga xabar yuborilmadi!\nBuyurtma #" + str(order_uid))
                    except:
                        pass

        cart_clear(tg_id)
        await state.clear()

    except Exception as e:
        logger.error(f"finalize_order xatosi: {e}")
        await message.answer(f"❌ Xatolik yuz berdi: {e}\nIltimos qayta urinib ko'ring.")
        await state.clear()

# --- ORDER CONFIRM/REJECT ---
@dp.callback_query(F.data.startswith("confirm_order_"))
async def confirm_order(callback: types.CallbackQuery):
    order_uid = callback.data.split("_", 2)[2]
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE orders SET status='confirmed', confirmed_at=%s WHERE order_uid=%s", (now_str(), order_uid))
    c.execute("SELECT * FROM orders WHERE order_uid=%s", (order_uid,))
    order = c.fetchone()
    conn.commit()
    conn.close()

    if order:
        await assign_courier(order_uid, order['user_tg_id'], order['shop_id'])

    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer("✅ Tasdiqlandi")

@dp.callback_query(F.data.startswith("reject_order_"))
async def reject_order(callback: types.CallbackQuery):
    order_uid = callback.data.split("_", 2)[2]
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE orders SET status='rejected', confirmed_at=%s WHERE order_uid=%s", (now_str(), order_uid))
    c.execute("SELECT user_tg_id, shop_id FROM orders WHERE order_uid=%s", (order_uid,))
    order = c.fetchone()
    conn.commit()
    conn.close()

    if order:
        conn2 = get_db()
        cur = conn2.cursor()
        cur.execute("SELECT owner_tg_id, phone FROM shops WHERE id=%s", (order['shop_id'],))
        shop = cur.fetchone()
        conn2.close()

        kb = InlineKeyboardBuilder()
        if shop and shop['phone']:
            kb.button(text="📞 Do'kon egasi bilan aloqa", url=f"https://t.me/+{shop['phone']}")

        try:
            await bot.send_message(order['user_tg_id'],
                                   "❌ Buyurtmangiz rad etildi.",
                                   reply_markup=kb.as_markup() if shop and shop['phone'] else None)
        except:
            pass

    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer("❌ Rad etildi")

# --- COURIER ASSIGNMENT ---
async def assign_courier(order_uid, user_tg_id, shop_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM couriers WHERE shop_id=%s AND is_blocked=0 AND is_available=1 ORDER BY queue_order ASC", (shop_id,))
    couriers = c.fetchall()
    conn.close()

    if not couriers:
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(admin_id, f"⚠️ #{order_uid} buyurtmasi uchun kuryer topilmadi!")
            except:
                pass
        return

    available = [cur for cur in couriers if not cur['is_busy']]
    if not available:
        for admin_id in ADMIN_IDS:
            try:
                await bot.send_message(admin_id, f"⚠️ #{order_uid} — barcha kuryerlar band!")
            except:
                pass
        return

    courier = available[0]
    await send_courier_offer(order_uid, courier, couriers, 0)

async def send_courier_offer(order_uid, courier, all_couriers, attempt):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM orders WHERE order_uid=%s", (order_uid,))
    order = c.fetchone()
    conn.close()

    if not order:
        return

    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Olaman", callback_data=f"take_order_{order_uid}_{courier['tg_id']}_{attempt}")
    kb.button(text="❌ Olmayman", callback_data=f"skip_order_{order_uid}_{courier['tg_id']}_{attempt}")
    kb.adjust(2)

    conn2 = get_db()
    c2 = conn2.cursor()
    c2.execute("SELECT * FROM users WHERE tg_id=%s", (order['user_tg_id'],))
    user = c2.fetchone()
    conn2.close()

    if order['user_tg_id'] == 0:
        mijoz_qator = f"👤 {order['address']}"
        manzil_qator = ""
    else:
        mijoz_qator = f"👤 Mijoz: {user['full_name'] if user else 'Noma\'lum'}\n📱 Tel: {user['phone'] if user else ''}"
        manzil_qator = f"\n📍 Manzil: {order['address']}"

    text = (f"🆕 Yangi buyurtma #{order_uid}\n\n"
            f"{mijoz_qator}{manzil_qator}\n"
            f"🛍️ {order['products']}\n"
            f"💰 {order['total_sum']:,.0f} so'm")

    try:
        await bot.send_message(courier['tg_id'], text, reply_markup=kb.as_markup())
        if order.get('latitude') and order.get('longitude'):
            await bot.send_location(courier['tg_id'], order['latitude'], order['longitude'])
    except:
        pass

@dp.callback_query(F.data.startswith("take_order_"))
async def take_order(callback: types.CallbackQuery):
    parts = callback.data.split("_")
    order_uid = parts[2]
    courier_tg = int(parts[3])

    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE orders SET courier_tg_id=%s, status='confirmed' WHERE order_uid=%s", (courier_tg, order_uid))
    c.execute("UPDATE couriers SET is_busy=1 WHERE tg_id=%s", (courier_tg,))
    c.execute("SELECT user_tg_id FROM orders WHERE order_uid=%s", (order_uid,))
    order = c.fetchone()
    c.execute("SELECT full_name, phone FROM couriers WHERE tg_id=%s", (courier_tg,))
    courier = c.fetchone()
    conn.commit()
    conn.close()

    if order and courier and order['user_tg_id'] != 0:
        try:
            await bot.send_message(order['user_tg_id'],
                                   f"✅ Buyurtmangiz tasdiqlandi!\n🚚 Kuryer: {courier['full_name']}\n📱 Tel: {courier['phone']}")
        except:
            pass

    kb2 = InlineKeyboardBuilder()
    kb2.button(text="🚗 Yo'lda", callback_data=f"c_onway_{order_uid}")
    kb2.button(text="⚠️ Hatolik", callback_data=f"c_error_{order_uid}")
    kb2.adjust(2)
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer(f"✅ #{order_uid} buyurtma qabul qilindi!\nYo'lga chiqqanda bosing 👇", reply_markup=kb2.as_markup())
    await callback.answer("✅ Buyurtma olindi!")

@dp.callback_query(F.data.startswith("skip_order_"))
async def skip_order(callback: types.CallbackQuery):
    parts = callback.data.split("_")
    order_uid = parts[2]
    courier_tg = int(parts[3])
    attempt = int(parts[4])

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT shop_id FROM orders WHERE order_uid=%s", (order_uid,))
    order = c.fetchone()
    conn.close()

    if order:
        conn2 = get_db()
        c2 = conn2.cursor()
        c2.execute("SELECT * FROM couriers WHERE shop_id=%s AND is_blocked=0 AND is_available=1 AND is_busy=0 ORDER BY queue_order ASC",
                   (order['shop_id'],))
        couriers = c2.fetchall()
        conn2.close()

        next_attempt = attempt + 1
        available = [cur for cur in couriers if cur['tg_id'] != courier_tg]

        if available and next_attempt < len(couriers) * 2:
            next_courier = available[next_attempt % len(available)]
            await send_courier_offer(order_uid, next_courier, couriers, next_attempt)
        else:
            for admin_id in ADMIN_IDS:
                try:
                    await bot.send_message(admin_id, f"⚠️ #{order_uid} — hamma kuryer rad etdi!")
                except:
                    pass

    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer("❌ O'tkazildi")

# --- COURIER PANEL ---
@dp.message(F.text == "📦 Buyurtmalarim")
async def courier_orders(message: types.Message):
    tg_id = message.from_user.id
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM orders WHERE courier_tg_id=%s AND status NOT IN ('delivered','rejected') ORDER BY created_at DESC", (tg_id,))
    orders = c.fetchall()
    conn.close()

    if not orders:
        await message.answer("📦 Faol buyurtmalar yo'q")
        return

    kb = InlineKeyboardBuilder()
    for o in orders:
        kb.button(text=f"#{o['order_uid']} — {o['address'][:20]}", callback_data=f"courier_order_{o['order_uid']}")
    kb.adjust(1)
    await message.answer("📦 Buyurtmalarim:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("courier_order_"))
async def courier_order_detail(callback: types.CallbackQuery):
    order_uid = callback.data.split("_", 2)[2]
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM orders WHERE order_uid=%s", (order_uid,))
    o = c.fetchone()
    conn.close()

    if not o:
        await callback.message.answer("Buyurtma topilmadi")
        return

    user = get_user(o['user_tg_id'])
    text = (f"📦 #{order_uid}\n"
            f"👤 {user['full_name'] if user else 'Noma\'lum'}\n"
            f"📱 {user['phone'] if user else ''}\n"
            f"📍 {o['address']}\n"
            f"🛍️ {o['products']}\n"
            f"💰 {o['total_sum']:,.0f} so'm\n"
            f"💳 {o['payment_type']}")

    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Qabul qildim", callback_data=f"c_accept_{order_uid}")
    kb.button(text="🚗 Yo'lda", callback_data=f"c_onway_{order_uid}")
    kb.button(text="✅ Yetib bordim", callback_data=f"c_delivered_{order_uid}")
    kb.button(text="⚠️ Hatolik", callback_data=f"c_error_{order_uid}")

    if o['latitude'] and o['longitude']:
        kb.button(text="🗺️ Google Maps", url=f"https://maps.google.com/?q={o['latitude']},{o['longitude']}")

    kb.adjust(2)
    await callback.message.answer(text, reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("c_accept_"))
async def courier_accept(callback: types.CallbackQuery):
    order_uid = callback.data.split("_", 2)[2]
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT user_tg_id FROM orders WHERE order_uid=%s", (order_uid,))
    o = c.fetchone()
    conn.close()
    if o and o['user_tg_id'] != 0:
        try:
            await bot.send_message(o['user_tg_id'], f"✅ Kuryer #{order_uid} buyurtmangizni qabul qildi!")
        except:
            pass
    await callback.answer("✅")

@dp.callback_query(F.data.startswith("c_onway_"))
async def courier_onway(callback: types.CallbackQuery):
    order_uid = callback.data.split("_", 2)[2]
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE orders SET status='on_way' WHERE order_uid=%s", (order_uid,))
    c.execute("SELECT * FROM orders WHERE order_uid=%s", (order_uid,))
    o = c.fetchone()
    conn.commit()
    conn.close()
    if o and o['user_tg_id'] != 0:
        try:
            await bot.send_message(o['user_tg_id'], f"🚗 Kuryer yo'lda! Buyurtma #{order_uid}")
        except:
            pass
    kb = InlineKeyboardBuilder()
    if o and o['latitude'] and o['longitude']:
        kb.button(text="🗺️ Google Maps", url=f"https://maps.google.com/?q={o['latitude']},{o['longitude']}")
    kb.button(text="✅ Yetkazildi!", callback_data=f"c_delivered_{order_uid}")
    kb.button(text="⚠️ Hatolik", callback_data=f"c_error_{order_uid}")
    kb.adjust(1)
    try:
        await callback.message.edit_reply_markup(reply_markup=kb.as_markup())
    except:
        pass
    await callback.answer("🚗 Yo'lda!")

@dp.callback_query(F.data.startswith("c_delivered_"))
async def courier_delivered(callback: types.CallbackQuery):
    order_uid = callback.data.split("_", 2)[2]
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE orders SET status='delivered', delivered_at=%s WHERE order_uid=%s", (now_str(), order_uid))
    c.execute("SELECT * FROM orders WHERE order_uid=%s", (order_uid,))
    o = c.fetchone()
    c.execute("UPDATE couriers SET is_busy=0 WHERE tg_id=%s", (callback.from_user.id,))
    conn.commit()
    conn.close()

    if o:
        conn2 = get_db()
        c2 = conn2.cursor()
        c2.execute("SELECT owner_tg_id, name FROM shops WHERE id=%s", (o['shop_id'],))
        shop = c2.fetchone()
        conn2.close()
        if shop:
            try:
                await bot.send_message(
                    shop['owner_tg_id'],
                    f"✅ Buyurtma #{order_uid} yetkazildi!\n"
                    f"🛍️ {o['products']}\n"
                    f"💰 {o['total_sum']:,.0f} so'm\n"
                    f"💳 {o['payment_type']}"
                )
            except:
                pass

    if o and o['user_tg_id'] != 0:
        kb = InlineKeyboardBuilder()
        for i in range(1, 6):
            kb.button(text=f"{'⭐' * i}", callback_data=f"rate_{order_uid}_{i}")
        kb.adjust(5)
        try:
            await bot.send_message(o['user_tg_id'],
                                   f"✅ Buyurtmangiz #{order_uid} yetkazildi!\n⭐ Iltimos, baho bering:",
                                   reply_markup=kb.as_markup())
        except:
            pass
        try:
            conn3 = get_db()
            c3 = conn3.cursor()
            c3.execute("SELECT * FROM shops WHERE id=%s", (o['shop_id'],))
            shop3 = c3.fetchone()
            conn3.close()
            user3 = get_user(o['user_tg_id'])
            pdf_buf, fmt = generate_pdf_receipt(dict(o), dict(user3) if user3 else {}, dict(shop3) if shop3 else {})
            if fmt == "pdf":
                await bot.send_document(o['user_tg_id'],
                    types.BufferedInputFile(pdf_buf.read(), filename=f"chek_{order_uid}.pdf"),
                    caption=f"🧾 Buyurtma #{order_uid} cheki")
            else:
                await bot.send_photo(o['user_tg_id'],
                    types.BufferedInputFile(pdf_buf.read(), filename=f"chek_{order_uid}.png"),
                    caption=f"🧾 Buyurtma #{order_uid} cheki")
        except Exception as e:
            logger.error(f"PDF chek xatosi: {e}")

    await callback.answer("✅ Yetkazildi!")
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer(f"✅ #{order_uid} buyurtma yetkazildi deb belgilandi!\n🆓 Endi bo'shsiz.")

@dp.callback_query(F.data.startswith("c_error_"))
async def courier_error(callback: types.CallbackQuery):
    order_uid = callback.data.split("_", 2)[2]
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT user_tg_id, shop_id FROM orders WHERE order_uid=%s", (order_uid,))
    o = c.fetchone()
    conn.close()

    if o:
        conn2 = get_db()
        cur = conn2.cursor()
        cur.execute("SELECT phone FROM shops WHERE id=%s", (o['shop_id'],))
        shop = cur.fetchone()
        conn2.close()

        kb = InlineKeyboardBuilder()
        if shop and shop['phone']:
            kb.button(text="📞 Do'kon egasi bilan aloqa", url=f"https://t.me/+{shop['phone']}")
        if o['user_tg_id'] != 0:
            try:
                await bot.send_message(o['user_tg_id'], f"⚠️ #{order_uid} buyurtmada muammo yuz berdi.",
                                       reply_markup=kb.as_markup() if shop and shop['phone'] else None)
            except:
                pass
    await callback.answer("⚠️")

# --- RATING ---
@dp.callback_query(F.data.startswith("rate_"))
async def rate_shop(callback: types.CallbackQuery):
    parts = callback.data.split("_")
    order_uid = parts[1]
    stars = int(parts[2])

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT shop_id FROM orders WHERE order_uid=%s", (order_uid,))
    o = c.fetchone()
    if o:
        c.execute("SELECT rating, rating_count FROM shops WHERE id=%s", (o['shop_id'],))
        shop = c.fetchone()
        new_count = shop['rating_count'] + 1
        new_rating = (shop['rating'] * shop['rating_count'] + stars) / new_count
        c.execute("UPDATE shops SET rating=%s, rating_count=%s WHERE id=%s", (new_rating, new_count, o['shop_id']))
        conn.commit()
    conn.close()

    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer(f"{'⭐' * stars} Rahmat!")

# --- COURIER BUSY ---
@dp.message(F.text == "📵 Bandman")
async def toggle_busy(message: types.Message):
    tg_id = message.from_user.id
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_available FROM couriers WHERE tg_id=%s", (tg_id,))
    courier = c.fetchone()
    if courier:
        new_val = 0 if courier['is_available'] else 1
        c.execute("UPDATE couriers SET is_available=%s WHERE tg_id=%s", (new_val, tg_id))
        conn.commit()
        status = "✅ Faol" if new_val else "📵 Band (buyurtma olmaysiz)"
        await message.answer(f"Holat o'zgartirildi: {status}")
    conn.close()

# --- COURIER REPORTS ---
@dp.message(F.text == "📅 Kunlik hisobot")
async def courier_daily(message: types.Message):
    tg_id = message.from_user.id
    today = datetime.now().strftime("%d.%m.%Y")
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM orders WHERE courier_tg_id=%s AND delivered_at LIKE %s", (tg_id, f"{today}%"))
    orders = c.fetchall()
    conn.close()

    count = len(orders)
    total = sum(o['total_sum'] for o in orders)
    await message.answer(f"📅 Bugungi hisobot\n\n📦 Buyurtmalar: {count}\n💰 Jami: {total:,.0f} so'm")

@dp.message(F.text == "📊 30 kunlik hisobot")
async def report_30(message: types.Message):
    tg_id = message.from_user.id
    role = user_role(tg_id)

    conn = get_db()
    c = conn.cursor()

    if role == "courier":
        c.execute("SELECT * FROM orders WHERE courier_tg_id=%s AND status='delivered'", (tg_id,))
        orders = c.fetchall()
        count = len(orders)
        total = sum(o['total_sum'] for o in orders)
        await message.answer(f"📊 30 kunlik hisobot\n\n📦 Yetkazilgan: {count}\n💰 Jami: {total:,.0f} so'm")

    elif role == "shop":
        shop = get_shop_by_owner(tg_id)
        if shop:
            c.execute("SELECT * FROM orders WHERE shop_id=%s AND status='delivered'", (shop['id'],))
            orders = c.fetchall()
            count = len(orders)
            total = sum(o['total_sum'] for o in orders)
            percent = shop['admin_percent']
            admin_share = total * percent / 100
            await message.answer(
                f"📊 30 kunlik hisobot\n\n📦 Buyurtmalar: {count}\n💰 Daromad: {total:,.0f} so'm\n"
                f"📊 Admin %: {percent}%\n💸 Admin ulushi: {admin_share:,.0f} so'm"
            )
    conn.close()

# --- SHOP PANEL ---
@dp.message(F.text == "🔓 Do'konni ochish/yopish")
async def toggle_shop(message: types.Message):
    shop = get_shop_by_owner(message.from_user.id)
    if not shop:
        return
    conn = get_db()
    c = conn.cursor()
    new_val = 0 if shop['is_open'] else 1
    c.execute("UPDATE shops SET is_open=%s WHERE id=%s", (new_val, shop['id']))
    conn.commit()
    conn.close()
    status = "🔓 Ochiq" if new_val else "🔒 Yopiq"
    await message.answer(f"Do'kon holati: {status}")

@dp.message(F.text == "✏️ Do'konni tahrirlash")
async def edit_shop_menu(message: types.Message):
    kb = InlineKeyboardBuilder()
    kb.button(text="🏪 Do'kon nomi", callback_data="edit_shop_name")
    kb.button(text="💳 Karta raqam", callback_data="edit_shop_card")
    kb.button(text="⏰ Ish vaqti", callback_data="edit_shop_time")
    kb.adjust(1)
    await message.answer("✏️ Nimani tahrirlash?", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("edit_shop_"))
async def edit_shop_field(callback: types.CallbackQuery, state: FSMContext):
    field = callback.data.split("_", 2)[2]
    await state.update_data(field=field)
    if field == "card":
        await callback.message.answer("💳 Yangi karta raqamini kiriting:\nMisol: 8600 1234 5678 9012")
        await state.set_state(ShopEditState.card)
    elif field == "time":
        await callback.message.answer("⏰ Yangi ish vaqtini kiriting:\nMisol: 07:00 - 23:00")
        await state.set_state(ShopEditState.name)
    else:
        await callback.message.answer("✍️ Yangi nomni kiriting:")
        await state.set_state(ShopEditState.name)

@dp.message(ShopEditState.card)
async def save_shop_card(message: types.Message, state: FSMContext):
    await state.update_data(card_number=message.text)
    await message.answer("👤 Karta egasining ismini kiriting:\nMisol: OLIMBEK TOSHMATOV")
    await state.set_state(ShopEditState.card_name)

@dp.message(ShopEditState.card_name)
async def save_shop_card_name(message: types.Message, state: FSMContext):
    data = await state.get_data()
    shop = get_shop_by_owner(message.from_user.id)
    if not shop:
        await state.clear()
        return

    card_info = f"{data['card_number']} | {message.text.upper()}"
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE shops SET card_number=%s WHERE id=%s", (card_info, shop['id']))
    conn.commit()
    conn.close()

    await message.answer(f"✅ Saqlandi!\n💳 Karta: {data['card_number']}\n👤 Ism: {message.text.upper()}")
    await state.clear()

@dp.message(ShopEditState.name)
async def save_shop_edit(message: types.Message, state: FSMContext):
    data = await state.get_data()
    field = data['field']
    shop = get_shop_by_owner(message.from_user.id)
    if not shop:
        await state.clear()
        return

    field_map = {"name": "name", "time": "work_time"}
    db_field = field_map.get(field, field)

    conn = get_db()
    c = conn.cursor()
    c.execute(f"UPDATE shops SET {db_field}=%s WHERE id=%s", (message.text, shop['id']))
    conn.commit()
    conn.close()

    await message.answer("✅ Saqlandi!")
    await state.clear()

# --- PRODUCTS ---
@dp.message(F.text == "📦 Mahsulotlar")
async def products_menu(message: types.Message):
    tg_id = message.from_user.id
    role = user_role(tg_id)

    if role == "shop":
        shop = get_shop_by_owner(tg_id)
    elif role == "admin":
        await message.answer("Admin: Qaysi do'kon mahsulotlarini ko'rmoqchisiz?")
        return
    else:
        return

    if not shop:
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM products WHERE shop_id=%s AND (is_available IS NULL OR is_available=1)", (shop['id'],))
    prods = c.fetchall()
    conn.close()

    kb = InlineKeyboardBuilder()
    for p in prods:
        kb.button(text=f"✏️ {p['name']} — {p['price']:,.0f}", callback_data=f"edit_prod_{p['id']}")
    kb.button(text="➕ Mahsulot qo'shish", callback_data=f"add_prod_{shop['id']}")
    kb.button(text="🔄 Fayl yangilash", callback_data=f"excel_update_{shop['id']}")
    kb.adjust(1)
    await message.answer(f"📦 Mahsulotlar ({len(prods)} ta):", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("add_prod_"))
async def add_product(callback: types.CallbackQuery, state: FSMContext):
    shop_id = int(callback.data.split("_")[2])
    await state.update_data(shop_id=shop_id)
    await callback.message.answer("✍️ Mahsulot nomini kiriting:")
    await state.set_state(AddProductState.name)

@dp.message(AddProductState.name)
async def add_prod_name(message: types.Message, state: FSMContext):
    await state.update_data(prod_name=message.text)
    await message.answer("💰 Narxini kiriting (so'm):")
    await state.set_state(AddProductState.price)

@dp.message(AddProductState.price)
async def add_prod_price(message: types.Message, state: FSMContext):
    try:
        price = float(message.text.replace(",", "").replace(" ", ""))
    except:
        await message.answer("❌ Noto'g'ri narx. Raqam kiriting:")
        return

    data = await state.get_data()
    uid = gen_id()
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO products (id, shop_id, name, price, created_at) VALUES (%s,%s,%s,%s,%s) ON CONFLICT (id) DO NOTHING",
              (uid, data['shop_id'], data['prod_name'], price, now_str()))
    conn.commit()
    conn.close()

    await message.answer(f"✅ {data['prod_name']} — {price:,.0f} so'm qo'shildi!")
    await state.clear()

@dp.callback_query(F.data.startswith("edit_prod_"))
async def edit_product(callback: types.CallbackQuery, state: FSMContext):
    prod_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM products WHERE id=%s", (prod_id,))
    prod = c.fetchone()
    conn.close()

    if not prod:
        await callback.answer("Topilmadi")
        return

    avail = prod['is_available'] if prod['is_available'] is not None else 1
    status_text = "✅ Mavjud" if avail else "❌ Tugagan"
    toggle_text = "❌ Tugadi deb belgilash" if avail else "✅ Mavjud deb belgilash"
    kb = InlineKeyboardBuilder()
    kb.button(text="✏️ Nomini o'zgartir", callback_data=f"ep_name_{prod_id}")
    kb.button(text="💰 Narxini o'zgartir", callback_data=f"ep_price_{prod_id}")
    kb.button(text=toggle_text, callback_data=f"ep_toggle_{prod_id}")
    kb.button(text="🗑️ O'chirish", callback_data=f"ep_del_{prod_id}")
    kb.adjust(1)
    await callback.message.answer(
        f"📦 {prod['name']} — {prod['price']:,.0f} so'm\nHolat: {status_text}\nNimani o'zgartirish?",
        reply_markup=kb.as_markup()
    )

@dp.callback_query(F.data.startswith("ep_name_") | F.data.startswith("ep_price_"))
async def edit_prod_field(callback: types.CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    field = parts[1]
    prod_id = int(parts[2])
    await state.update_data(prod_id=prod_id, field=field)
    await callback.message.answer("✍️ Yangi qiymatni kiriting:")
    await state.set_state(EditProductState.name)

@dp.message(EditProductState.name)
async def save_prod_edit(message: types.Message, state: FSMContext):
    data = await state.get_data()
    conn = get_db()
    c = conn.cursor()
    if data['field'] == 'name':
        c.execute("UPDATE products SET name=%s WHERE id=%s", (message.text, data['prod_id']))
    else:
        try:
            price = float(message.text.replace(",", "").replace(" ", ""))
            c.execute("UPDATE products SET price=%s WHERE id=%s", (price, data['prod_id']))
        except:
            await message.answer("❌ Noto'g'ri narx")
            conn.close()
            return
    conn.commit()
    conn.close()
    await message.answer("✅ Saqlandi!")
    await state.clear()

@dp.callback_query(F.data.startswith("ep_del_"))
async def delete_product(callback: types.CallbackQuery):
    prod_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM products WHERE id=%s", (prod_id,))
    conn.commit()
    conn.close()
    await callback.message.answer("🗑️ Mahsulot o'chirildi")

@dp.callback_query(F.data.startswith("ep_toggle_"))
async def toggle_product_availability(callback: types.CallbackQuery):
    prod_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name, is_available FROM products WHERE id=%s", (prod_id,))
    prod = c.fetchone()
    if not prod:
        await callback.answer("Topilmadi")
        conn.close()
        return
    new_val = 0 if (prod['is_available'] if prod['is_available'] is not None else 1) else 1
    c.execute("UPDATE products SET is_available=%s WHERE id=%s", (new_val, prod_id))
    conn.commit()
    conn.close()
    status = "✅ Mavjud" if new_val else "❌ Tugagan"
    await callback.message.answer(f"'{prod['name']}' holati: {status}")
    await callback.answer()

# --- EXCEL UPLOAD ---
@dp.callback_query(F.data.startswith("excel_update_"))
async def excel_update(callback: types.CallbackQuery, state: FSMContext):
    shop_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM products WHERE shop_id=%s", (shop_id,))
    prods = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mahsulotlar"

    header_fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["№", "Mahsulot nomi", "Narxi (so'm)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18

    if prods:
        for i, p in enumerate(prods, 1):
            ws.append([i, p['name'], p['price']])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        await callback.message.answer_document(
            types.BufferedInputFile(buf.read(), filename="mahsulotlar.xlsx"),
            caption="📊 Hozirgi mahsulotlar fayli!\n\n✏️ Faylni o'zgartiring va qayta yuboring."
        )
    else:
        examples = [[1, "Osh", 25000], [2, "Shashlik", 35000], [3, "Lag'mon", 22000]]
        for row in examples:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        await callback.message.answer_document(
            types.BufferedInputFile(buf.read(), filename="shablon.xlsx"),
            caption="📋 Shablon faylni yuklab oling, to'ldiring va qayta yuboring."
        )

    await state.update_data(shop_id=shop_id, update_mode=True)
    await state.set_state(ExcelUpdateState.waiting)

@dp.message(ExcelUpdateState.waiting, F.document)
async def excel_update_file(message: types.Message, state: FSMContext):
    data = await state.get_data()
    shop_id = data['shop_id']

    file = await bot.get_file(message.document.file_id)
    file_bytes = await bot.download_file(file.file_path)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes.read()))
        ws = wb.active
        conn = get_db()
        c = conn.cursor()
        c.execute("DELETE FROM products WHERE shop_id=%s", (shop_id,))
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[2]:
                try:
                    price = float(str(row[2]).replace(",", "").replace(" ", ""))
                    uid = gen_id()
                    c.execute("INSERT INTO products (id, shop_id, name, price, created_at) VALUES (%s,%s,%s,%s,%s)",
                              (uid, shop_id, str(row[1]), price, now_str()))
                    count += 1
                except:
                    pass
        conn.commit()
        conn.close()
        await message.answer(f"✅ {count} ta mahsulot yangilandi!")
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}")
    await state.clear()

# --- SHOP: KURYER QO'SHISH ---
@dp.message(F.text == "🚚 Kuryer qo'shish")
async def shop_add_courier_start(message: types.Message, state: FSMContext):
    shop = get_shop_by_owner(message.from_user.id)
    if not shop:
        return
    await message.answer(
        "🆔 Kuryer Telegram ID sini kiriting:\n\n"
        "📌 Kuryer o'z ID sini bilmasa:\n"
        "→ @userinfobot ga /start yubortiring"
    )
    await state.set_state(ShopAddCourierState.tg_id)

@dp.message(ShopAddCourierState.tg_id)
async def shop_add_courier_tg_id(message: types.Message, state: FSMContext):
    try:
        tg_id = int(message.text.strip())
    except:
        await message.answer("❌ Noto'g'ri ID! Faqat raqam kiriting.")
        return
    await state.update_data(tg_id=tg_id)
    await message.answer("👤 Kuryer ism familyasini kiriting:")
    await state.set_state(ShopAddCourierState.name)

@dp.message(ShopAddCourierState.name)
async def shop_add_courier_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text.strip())
    await message.answer("📱 Kuryer telefon raqamini kiriting:")
    await state.set_state(ShopAddCourierState.phone)

@dp.message(ShopAddCourierState.phone)
async def shop_add_courier_phone(message: types.Message, state: FSMContext):
    data = await state.get_data()
    shop = get_shop_by_owner(message.from_user.id)
    if not shop:
        await state.clear()
        return

    tg_id = data['tg_id']
    name = data['name']
    phone = message.text.strip()
    uid = gen_id()

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) as cnt FROM couriers WHERE shop_id=%s", (shop['id'],))
    queue_order = c.fetchone()['cnt']

    try:
        c.execute(
            "INSERT INTO couriers (id, tg_id, full_name, phone, shop_id, queue_order, registered_at) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (uid, tg_id, name, phone, shop['id'], queue_order, now_str())
        )
        conn.commit()
        conn.close()

        await message.answer(
            f"✅ Kuryer qo'shildi!\n\n"
            f"🆔 TG ID: {tg_id}\n"
            f"👤 Ism: {name}\n"
            f"📱 Tel: {phone}\n"
            f"🏪 Do'kon: {shop['name']}"
        )
        try:
            await bot.send_message(
                tg_id,
                f"✅ Siz {shop['name']} do'koniga kuryer sifatida qo'shildingiz!\n"
                f"Botni ishlatish uchun /start bosing."
            )
        except:
            await message.answer("⚠️ Kuryerga xabar yuborilmadi (botni boshlamagandirmi)")
    except Exception as e:
        conn.close()
        await message.answer(f"❌ Xatolik: Bu TG ID allaqachon ro'yxatda bo'lishi mumkin.\n{e}")

    await state.clear()

# --- VACATION ---
@dp.message(F.text == "🏖️ Ta'tilda")
async def vacation_start(message: types.Message, state: FSMContext):
    await message.answer("📅 Ta'til tugash sanasini kiriting (DD.MM.YYYY):", reply_markup=back_kb())
    await state.set_state(VacationState.end_date)

@dp.message(VacationState.end_date)
async def vacation_set(message: types.Message, state: FSMContext):
    if message.text == "⬅️ Orqaga":
        await state.clear()
        await message.answer("Bosh menyu:", reply_markup=shop_main_kb())
        return

    try:
        end_date = datetime.strptime(message.text, "%d.%m.%Y")
        if end_date.date() < date.today():
            await message.answer("❌ O'tgan sana. Qaytadan kiriting:")
            return
    except:
        await message.answer("❌ Noto'g'ri format. Misol: 05.02.2026")
        return

    shop = get_shop_by_owner(message.from_user.id)
    if shop:
        conn = get_db()
        c = conn.cursor()
        c.execute("UPDATE shops SET vacation_until=%s, is_open=0 WHERE id=%s", (message.text, shop['id']))
        conn.commit()
        conn.close()

    await message.answer(f"🏖️ Ta'til belgilandi: {message.text} gacha\nDo'kon yopildi.")
    await state.clear()

# --- PHONE ORDER ---
@dp.message(F.text == "📞 Telefon buyurtma")
async def phone_order_menu(message: types.Message):
    shop = get_shop_by_owner(message.from_user.id)
    if not shop:
        return

    kb = InlineKeyboardBuilder()
    kb.button(text="📱 Nomer qo'shish/tahrirlash", callback_data="edit_shop_phone")
    kb.button(text="📝 Buyurtma kiritish", callback_data=f"new_phone_order_{shop['id']}")
    kb.adjust(1)
    await message.answer("📞 Telefon orqali buyurtma:", reply_markup=kb.as_markup())

@dp.callback_query(F.data == "edit_shop_phone")
async def edit_shop_phone(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.answer("📱 Yangi telefon raqamni kiriting:")
    await state.set_state(AddPhoneState.phone)

@dp.message(AddPhoneState.phone)
async def save_shop_phone(message: types.Message, state: FSMContext):
    shop = get_shop_by_owner(message.from_user.id)
    if shop:
        conn = get_db()
        c = conn.cursor()
        c.execute("UPDATE shops SET phone=%s WHERE id=%s", (message.text, shop['id']))
        conn.commit()
        conn.close()
    await message.answer("✅ Nomer saqlandi!")
    await state.clear()

@dp.callback_query(F.data.startswith("new_phone_order_"))
async def new_phone_order(callback: types.CallbackQuery, state: FSMContext):
    shop_id = int(callback.data.split("_")[3])
    await state.update_data(shop_id=shop_id)
    await callback.message.answer("1️⃣ Mijoz ism familyasini kiriting:")
    await state.set_state(PhoneOrderState.client_name)

@dp.message(PhoneOrderState.client_name)
async def po_name(message: types.Message, state: FSMContext):
    await state.update_data(client_name=message.text)
    await message.answer("2️⃣ Mijoz telefon raqamini kiriting:")
    await state.set_state(PhoneOrderState.client_phone)

@dp.message(PhoneOrderState.client_phone)
async def po_phone(message: types.Message, state: FSMContext):
    await state.update_data(client_phone=message.text)
    await message.answer("3️⃣ Mijoz manzilini kiriting:")
    await state.set_state(PhoneOrderState.address)

@dp.message(PhoneOrderState.address)
async def po_address(message: types.Message, state: FSMContext):
    await state.update_data(address=message.text)
    await message.answer("4️⃣ Mahsulotni kiriting:")
    await state.set_state(PhoneOrderState.product)

@dp.message(PhoneOrderState.product)
async def po_product(message: types.Message, state: FSMContext):
    await state.update_data(product=message.text)
    await message.answer("5️⃣ Narxini kiriting (so'm):")
    await state.set_state(PhoneOrderState.price)

@dp.message(PhoneOrderState.price)
async def po_price(message: types.Message, state: FSMContext):
    try:
        price = float(message.text.replace(",", "").replace(" ", ""))
    except:
        await message.answer("❌ Noto'g'ri narx. Faqat raqam kiriting.")
        return

    await state.update_data(price=price)
    kb = InlineKeyboardBuilder()
    kb.button(text="💵 Naqd", callback_data="po_pay_cash")
    kb.button(text="💳 Karta", callback_data="po_pay_card")
    kb.adjust(2)
    await message.answer("6️⃣ To'lov turini tanlang:", reply_markup=kb.as_markup())
    await state.set_state(PhoneOrderState.payment)

@dp.callback_query(F.data.in_(["po_pay_cash", "po_pay_card"]), StateFilter(PhoneOrderState.payment))
async def po_payment(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    payment = "Naqd" if callback.data == "po_pay_cash" else "Karta"
    order_uid = str(gen_id())
    shop_id = data['shop_id']
    price = data.get('price', 0)
    address = data.get('address', '')

    conn = get_db()
    c = conn.cursor()
    order_id2 = gen_id()
    c.execute('''INSERT INTO orders (id, order_uid, user_tg_id, shop_id, products, total_sum, address,
              payment_type, status, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)''',
              (order_id2, order_uid, 0, shop_id, data['product'], price,
               f"Tel: {data['client_name']} / {data['client_phone']} / {address}",
               payment, 'confirmed', now_str()))
    conn.commit()
    conn.close()

    await callback.message.answer(
        f"✅ Telefon buyurtma kiritildi!\n"
        f"📦 #{order_uid}\n"
        f"👤 {data['client_name']}\n"
        f"📱 {data['client_phone']}\n"
        f"📍 {address}\n"
        f"🛍️ {data['product']}\n"
        f"💰 {price:,.0f} so'm\n"
        f"💳 {payment}"
    )
    await state.clear()
    await assign_courier(order_uid, 0, shop_id)

# --- SHOP REPORTS ---
@dp.message(F.text == "📅 Bugungi hisobot")
async def daily_report(message: types.Message):
    tg_id = message.from_user.id
    role = user_role(tg_id)
    today = datetime.now().strftime("%d.%m.%Y")

    conn = get_db()
    c = conn.cursor()

    if role == "shop":
        shop = get_shop_by_owner(tg_id)
        if shop:
            c.execute("SELECT * FROM orders WHERE shop_id=%s AND created_at LIKE %s AND status='delivered'",
                      (shop['id'], f"{today}%"))
            orders = c.fetchall()
            count = len(orders)
            total = sum(o['total_sum'] for o in orders)
            await message.answer(f"📅 Bugungi hisobot ({today})\n\n📦 Buyurtmalar: {count}\n💰 Daromad: {total:,.0f} so'm")
    elif role == "admin":
        c.execute("SELECT * FROM orders WHERE created_at LIKE %s AND status='delivered'", (f"{today}%",))
        orders = c.fetchall()
        count = len(orders)
        total = sum(o['total_sum'] for o in orders)
        await message.answer(f"📅 Bugungi umumiy hisobot ({today})\n\n📦 Buyurtmalar: {count}\n💰 Daromad: {total:,.0f} so'm")

    conn.close()

# --- CHAT TIZIMI ---
def chat_set(tg_id, partner, chat_type, order_id=""):
    conn = get_db()
    c = conn.cursor()
    c.execute('''INSERT INTO active_sessions (tg_id, partner, chat_type, order_id)
                 VALUES (%s,%s,%s,%s)
                 ON CONFLICT(tg_id) DO UPDATE SET
                 partner=excluded.partner, chat_type=excluded.chat_type, order_id=excluded.order_id''',
              (tg_id, partner, chat_type, order_id))
    conn.commit()
    conn.close()

def chat_get_session(tg_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM active_sessions WHERE tg_id=%s", (tg_id,))
    row = c.fetchone()
    conn.close()
    if not row:
        return None
    return {"partner": row["partner"], "type": row["chat_type"], "order": row["order_id"]}

def chat_remove(tg_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM active_sessions WHERE tg_id=%s", (tg_id,))
    conn.commit()
    conn.close()

def chat_clear_all():
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM active_sessions")
    conn.commit()
    conn.close()

def chat_end_kb():
    kb = ReplyKeyboardBuilder()
    kb.button(text="🔴 Chatni tugatish")
    return kb.as_markup(resize_keyboard=True, one_time_keyboard=False)

async def notify_partner_chat_started(partner_tg_id, initiator_tg_id, from_label, chat_type, order_uid=""):
    order_info = (f" (Buyurtma #{order_uid})") if order_uid else ""
    kb = InlineKeyboardBuilder()
    kb.button(text="💬 Javob berish", callback_data=f"reply_chat:{initiator_tg_id}:{chat_type}")
    msg = f"💬 {from_label} siz bilan gaplashmoqchi{order_info}.\n\nJavob berish uchun tugmani bosing:"
    try:
        await bot.send_message(partner_tg_id, msg, reply_markup=kb.as_markup())
    except Exception as e:
        logger.error(f"Chat notify error: {e}")

@dp.callback_query(F.data.startswith("chat_shop_"))
async def start_chat_with_shop(callback: types.CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    shop_id = int(parts[2])
    order_uid = parts[3]

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT owner_tg_id, name FROM shops WHERE id=%s", (shop_id,))
    shop = c.fetchone()
    conn.close()

    if not shop:
        await callback.message.answer("Do'kon topilmadi")
        return

    user_tg = callback.from_user.id
    shop_tg = shop["owner_tg_id"]

    chat_set(user_tg, shop_tg, "user_shop", order_uid)
    chat_set(shop_tg, user_tg, "shop_user", order_uid)

    await state.set_state(ChatState.chatting)
    await callback.message.answer(
        f"💬 {shop['name']} egasi bilan chat boshlandi (Buyurtma #{order_uid}).\nYozing:",
        reply_markup=chat_end_kb()
    )
    user = get_user(user_tg)
    user_name = user["full_name"] if user else str(user_tg)
    await notify_partner_chat_started(shop_tg, user_tg, f"Mijoz {user_name}", "shop_user", order_uid)

@dp.message(ChatState.chatting)
async def chat_message(message: types.Message, state: FSMContext):
    tg_id = message.from_user.id
    text = message.text or ""

    if text == "🔴 Chatni tugatish":
        chat = chat_get_session(tg_id)
        if chat:
            partner = chat["partner"]
            chat_remove(tg_id)
            chat_remove(partner)
            try:
                await bot.send_message(partner, "🔴 Chat yakunlandi.", reply_markup=main_menu_kb(partner))
            except:
                pass
        await state.clear()
        await message.answer("✅ Chat yakunlandi.", reply_markup=main_menu_kb(tg_id))
        return

    chat = chat_get_session(tg_id)
    if not chat:
        await state.clear()
        await message.answer("Chat topilmadi.", reply_markup=main_menu_kb(tg_id))
        return

    partner = chat["partner"]
    role = user_role(tg_id)
    if role == "courier":
        courier = get_courier_by_tg(tg_id)
        sender_name = f"🚚 Kuryer {courier['full_name'] if courier else tg_id}"
    elif role == "shop":
        shop = get_shop_by_owner(tg_id)
        sender_name = f"🏪 {shop['name'] + ' egasi' if shop else 'Do\'kon egasi'}"
    elif role == "admin":
        sender_name = "👨‍💼 Admin"
    else:
        user = get_user(tg_id)
        sender_name = f"👤 {user['full_name'] if user else 'Mijoz'}"

    try:
        await bot.send_message(partner, f"💬 {sender_name}:\n{text}")
    except Exception as e:
        logger.error(f"Chat send error: {e}")
        await message.answer("❌ Xabar yuborilmadi.")
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO chats (from_tg_id, to_tg_id, message, chat_type, created_at, order_id) VALUES (%s,%s,%s,%s,%s,%s)",
              (tg_id, partner, text, chat["type"], now_str(), chat.get("order", "")))
    conn.commit()
    conn.close()

@dp.callback_query(F.data.startswith("reply_chat:"))
async def reply_chat_start(callback: types.CallbackQuery, state: FSMContext):
    parts = callback.data.split(":")
    if len(parts) < 3:
        await callback.answer("Xato")
        return
    try:
        partner_tg_id = int(parts[1])
        chat_type = parts[2]
    except ValueError:
        await callback.answer("Xato ID")
        return
    tg_id = callback.from_user.id

    existing_partner = chat_get_session(partner_tg_id)
    if existing_partner:
        chat_set(tg_id, partner_tg_id, chat_type, existing_partner.get("order", ""))
    else:
        chat_set(tg_id, partner_tg_id, chat_type, "")
        chat_set(partner_tg_id, tg_id, chat_type, "")

    await state.set_state(ChatState.chatting)
    await callback.message.answer("💬 Chat ochildi! Yozing:", reply_markup=chat_end_kb())
    await callback.answer()
    try:
        await bot.send_message(partner_tg_id, "💬 Javob keldi! Endi yozishingiz mumkin.")
    except:
        pass

@dp.message(F.text == "💬 Do'kon egasi")
async def user_or_courier_chat_shop(message: types.Message, state: FSMContext):
    tg_id = message.from_user.id
    role = user_role(tg_id)

    if role == "courier":
        courier = get_courier_by_tg(tg_id)
        if not courier:
            return
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT owner_tg_id, name FROM shops WHERE id=%s", (courier["shop_id"],))
        shop = c.fetchone()
        conn.close()
        if not shop:
            await message.answer("Do'kon topilmadi")
            return
        chat_set(tg_id, shop["owner_tg_id"], "courier_shop", "")
        chat_set(shop["owner_tg_id"], tg_id, "shop_courier", "")
        await state.set_state(ChatState.chatting)
        await message.answer(f"💬 {shop['name']} egasi bilan chat boshlandi.\nYozing:", reply_markup=chat_end_kb())
        await notify_partner_chat_started(shop["owner_tg_id"], tg_id, f"Kuryer {courier['full_name']}", "shop_courier")
    else:
        conn = get_db()
        c = conn.cursor()
        c.execute("SELECT shop_id FROM orders WHERE user_tg_id=%s ORDER BY created_at DESC LIMIT 1", (tg_id,))
        last_order = c.fetchone()
        if not last_order:
            await message.answer("❌ Hali buyurtma yo'q.")
            conn.close()
            return
        c.execute("SELECT owner_tg_id, name FROM shops WHERE id=%s", (last_order["shop_id"],))
        shop = c.fetchone()
        conn.close()
        if not shop:
            await message.answer("Do'kon topilmadi")
            return
        chat_set(tg_id, shop["owner_tg_id"], "user_shop", "")
        chat_set(shop["owner_tg_id"], tg_id, "shop_user", "")
        await state.set_state(ChatState.chatting)
        user = get_user(tg_id)
        await message.answer(f"💬 {shop['name']} egasi bilan chat boshlandi.\nYozing:", reply_markup=chat_end_kb())
        await notify_partner_chat_started(shop["owner_tg_id"], tg_id, f"Mijoz {user['full_name'] if user else tg_id}", "shop_user")

@dp.message(F.text == "💬 Admin")
async def user_or_courier_chat_admin(message: types.Message, state: FSMContext):
    tg_id = message.from_user.id
    admin_id = ADMIN_IDS[0]
    role = user_role(tg_id)

    chat_set(tg_id, admin_id, "user_admin", "")
    chat_set(admin_id, tg_id, "admin_user", "")
    await state.set_state(ChatState.chatting)
    await message.answer("💬 Admin bilan chat boshlandi.\nYozing:", reply_markup=chat_end_kb())

    if role == "courier":
        courier = get_courier_by_tg(tg_id)
        label = f"Kuryer {courier['full_name'] if courier else tg_id}"
    elif role == "shop":
        shop = get_shop_by_owner(tg_id)
        label = f"Do'kon egasi {shop['name'] if shop else tg_id}"
    else:
        user = get_user(tg_id)
        label = f"Mijoz {user['full_name'] if user else tg_id}"
    await notify_partner_chat_started(admin_id, tg_id, label, "admin_user")

# --- TOP MIJOZLAR ---
@dp.message(F.text == "🏆 Top mijozlar")
async def top_buyers(message: types.Message):
    if not is_admin(message.from_user.id):
        return
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT o.user_tg_id, u.full_name, u.phone,
               COUNT(o.id) as order_count,
               SUM(o.total_sum) as total_spent
        FROM orders o
        LEFT JOIN users u ON u.tg_id = o.user_tg_id
        WHERE o.status = 'delivered'
        GROUP BY o.user_tg_id, u.full_name, u.phone
        ORDER BY order_count DESC
        LIMIT 10
    """)
    rows = c.fetchall()
    conn.close()

    if not rows:
        await message.answer("Hali yetkazilgan buyurtma yo'q")
        return

    text = "🏆 <b>Eng ko'p buyurtma qilgan mijozlar:</b>\n\n"
    medals = ["🥇", "🥈", "🥉"]
    for i, r in enumerate(rows):
        medal = medals[i] if i < 3 else f"{i+1}."
        name = r['full_name'] or "Noma'lum"
        phone = r['phone'] or ""
        spent = r['total_spent'] or 0
        text += f"{medal} {name} ({phone})\n"
        text += f"   📦 {r['order_count']} ta buyurtma | 💰 {spent:,.0f} so'm\n\n"

    await message.answer(text, parse_mode="HTML")

# ===================== ADMIN PANEL =====================

@dp.message(F.text == "📱 Buyurtma berish")
async def admin_order_start(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM shops WHERE is_open=1 ORDER BY name")
    shops = c.fetchall()
    conn.close()
    if not shops:
        await message.answer("Hozir ochiq do'kon yo'q")
        return
    kb = InlineKeyboardBuilder()
    for s in shops:
        kb.button(text=f"🏪 {s['name']}", callback_data=f"adm_ord_shop_{s['id']}")
    kb.adjust(1)
    await message.answer("🏪 Qaysi do'konga buyurtma?", reply_markup=kb.as_markup())
    await state.set_state(AdminOrderState.choosing_shop)

@dp.callback_query(F.data.startswith("adm_ord_shop_"), StateFilter(AdminOrderState.choosing_shop))
async def admin_order_shop(callback: types.CallbackQuery, state: FSMContext):
    shop_id = int(callback.data.split("_")[3])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM shops WHERE id=%s", (shop_id,))
    shop = c.fetchone()
    conn.close()
    await state.update_data(shop_id=shop_id, shop_name=shop['name'])
    await callback.message.answer("📱 Mijoz telefon raqami:")
    await state.set_state(AdminOrderState.phone)
    await callback.answer()

@dp.message(AdminOrderState.phone)
async def admin_order_phone(message: types.Message, state: FSMContext):
    await state.update_data(client_phone=message.text)
    await message.answer("👤 Mijoz ismi:")
    await state.set_state(AdminOrderState.name)

@dp.message(AdminOrderState.name)
async def admin_order_name(message: types.Message, state: FSMContext):
    await state.update_data(client_name=message.text)
    await message.answer("📍 Manzil:")
    await state.set_state(AdminOrderState.address)

@dp.message(AdminOrderState.address)
async def admin_order_address(message: types.Message, state: FSMContext):
    await state.update_data(address=message.text)
    await message.answer("🛍️ Mahsulotlarni yozing:")
    await state.set_state(AdminOrderState.product)

@dp.message(AdminOrderState.product)
async def admin_order_product(message: types.Message, state: FSMContext):
    data = await state.get_data()
    await state.update_data(products=message.text)
    uid = gen_id()
    await state.update_data(order_uid=uid)

    text = (f"📋 Buyurtma:\n\n"
            f"🏪 Do'kon: {data['shop_name']}\n"
            f"📱 Tel: {data['client_phone']}\n"
            f"👤 Ism: {data['client_name']}\n"
            f"📍 Manzil: {data['address']}\n"
            f"🛍️ Mahsulot: {message.text}\n"
            f"🆔 #{uid}")

    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Yuborish", callback_data="adm_ord_confirm")
    kb.button(text="❌ Bekor qilish", callback_data="adm_ord_cancel")
    kb.adjust(2)
    await message.answer(text, reply_markup=kb.as_markup())
    await state.set_state(AdminOrderState.confirm)

@dp.callback_query(F.data == "adm_ord_cancel", StateFilter(AdminOrderState.confirm))
async def admin_order_cancel(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.answer("❌ Bekor qilindi", reply_markup=admin_main_kb())
    await callback.answer()

@dp.callback_query(F.data == "adm_ord_confirm", StateFilter(AdminOrderState.confirm))
async def admin_order_confirm(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    await state.clear()

    shop_id = data['shop_id']
    uid = data['order_uid']
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM shops WHERE id=%s", (shop_id,))
    shop = c.fetchone()
    c.execute("""INSERT INTO orders (order_uid, user_tg_id, shop_id, products, total_sum,
              payment_type, address, status, created_at, source)
              VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
              (uid, 0, shop_id, data['products'], 0,
               'naqd', data['address'], 'pending', now_str(), 'admin'))
    conn.commit()
    conn.close()

    log_admin_action(callback.from_user.id, "Buyurtma berdi", f"Do'kon: {data['shop_name']}, ID: #{uid}")

    shop_text = (f"📱 ADMINdan buyurtma!\n\n"
                 f"🆔 #{uid}\n"
                 f"👤 Mijoz: {data['client_name']}\n"
                 f"📱 Tel: {data['client_phone']}\n"
                 f"📍 Manzil: {data['address']}\n"
                 f"🛍️ Mahsulot: {data['products']}\n"
                 f"📅 Vaqt: {now_str()}")
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Tasdiqlash", callback_data=f"confirm_order_{uid}")
    kb.button(text="❌ Rad etish", callback_data=f"reject_order_{uid}")
    kb.adjust(2)
    if shop:
        try:
            await bot.send_message(shop['owner_tg_id'], shop_text, reply_markup=kb.as_markup())
        except:
            pass

    await callback.message.answer(f"✅ Buyurtma #{uid} do'konga yuborildi!", reply_markup=admin_main_kb())
    await callback.answer()

# --- HATOLIK ---
@dp.message(F.text == "⚠️ Hatolik")
async def user_report_error(message: types.Message, state: FSMContext):
    user = get_user(message.from_user.id)
    if not user:
        return
    await message.answer("📝 Qanday hatolik yuz berdi? Batafsil yozing:")
    await state.set_state(ReportErrorState.waiting_text)

@dp.message(ReportErrorState.waiting_text)
async def user_error_text(message: types.Message, state: FSMContext):
    await state.clear()
    user = get_user(message.from_user.id)

    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO chats (from_tg_id, to_tg_id, message, chat_type, created_at) VALUES (%s,%s,%s,%s,%s)",
              (message.from_user.id, 0, f"[HATOLIK] {message.text}", "error_report", now_str()))
    conn.commit()
    conn.close()

    admin_text = (f"⚠️ HATOLIK XABARI\n\n"
                  f"👤 {user['full_name'] if user else 'Noma\'lum'}\n"
                  f"📱 {user['phone'] if user else ''}\n"
                  f"🆔 TG ID: {message.from_user.id}\n"
                  f"📅 {now_str()}\n\n"
                  f"💬 Xabar:\n{message.text}")
    kb = InlineKeyboardBuilder()
    kb.button(text="💬 Mijoz bilan chat", callback_data=f"admin_chat_user_{message.from_user.id}")
    kb.button(text="✅ Ko'rib chiqildi", callback_data=f"error_done_{message.from_user.id}")
    kb.adjust(1)

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, admin_text, reply_markup=kb.as_markup())
        except:
            pass

    await message.answer("✅ Xabatingiz adminlarga yuborildi!")

@dp.callback_query(F.data.startswith("error_done_"))
async def error_done(callback: types.CallbackQuery):
    user_tg_id = int(callback.data.split("_")[2])
    await callback.message.edit_text(callback.message.text + "\n\n✅ Ko'rib chiqildi")
    try:
        await bot.send_message(user_tg_id, "✅ Hatoligingiz ko'rib chiqildi. Rahmat!")
    except:
        pass
    await callback.answer("✅ Bajarildi")

# --- ADMIN: DO'KONLAR ---
@dp.message(F.text == "🏪 Do'konlar boshqaruv")
async def admin_shops(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM shops ORDER BY created_at DESC")
    shops = c.fetchall()
    conn.close()

    kb = InlineKeyboardBuilder()
    for s in shops:
        status = "🟢" if s['is_open'] else "🔴"
        kb.button(text=f"{status} {s['name']} ⭐{s['rating']:.1f}", callback_data=f"admin_shop_{s['id']}")
    kb.button(text="➕ Yangi do'kon", callback_data="add_new_shop")
    kb.adjust(1)
    await message.answer("🏪 Do'konlar:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("admin_shop_"))
async def admin_shop_detail(callback: types.CallbackQuery):
    shop_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM shops WHERE id=%s", (shop_id,))
    shop = c.fetchone()
    c.execute("SELECT COUNT(*) as cnt FROM orders WHERE shop_id=%s", (shop_id,))
    order_count = c.fetchone()['cnt']
    conn.close()

    if not shop:
        await callback.message.answer("Topilmadi")
        return

    text = (f"🏪 {shop['name']}\n"
            f"🆔 ID: {shop['id']}\n"
            f"👤 Egasi TG: {shop['owner_tg_id']}\n"
            f"📱 Tel: {shop['phone'] or 'Yo\'q'}\n"
            f"💳 Karta: {shop['card_number'] or 'Yo\'q'}\n"
            f"⏰ Ish vaqti: {shop['work_time'] or 'Yo\'q'}\n"
            f"⭐ Reyting: {shop['rating']:.1f}\n"
            f"📦 Jami buyurtmalar: {order_count}\n"
            f"📊 Admin %: {shop['admin_percent']}%\n"
            f"📅 Qo'shilgan: {shop['created_at']}\n"
            f"Holat: {'🟢 Ochiq' if shop['is_open'] else '🔴 Yopiq'}")

    kb = InlineKeyboardBuilder()
    kb.button(text="📦 Buyurtmalar", callback_data=f"ashop_orders_{shop_id}")
    kb.button(text="✏️ Nom tahrirlash", callback_data=f"ashop_edit_name_{shop_id}")
    kb.button(text="💳 Karta tahrirlash", callback_data=f"ashop_edit_card_{shop_id}")
    kb.button(text="📱 Nomer tahrirlash", callback_data=f"ashop_edit_phone_{shop_id}")
    kb.button(text="📊 Admin % belgilash", callback_data=f"ashop_percent_{shop_id}")
    kb.button(text="🔓/🔒 Ochish/Yopish", callback_data=f"ashop_toggle_{shop_id}")
    kb.button(text="🗑️ O'chirish", callback_data=f"ashop_delete_{shop_id}")
    kb.button(text="⬅️ Orqaga", callback_data="back_admin_shops")
    kb.adjust(2)
    await callback.message.answer(text, reply_markup=kb.as_markup())

@dp.callback_query(F.data == "back_admin_shops")
async def back_admin_shops(callback: types.CallbackQuery):
    await callback.message.delete()
    await admin_shops(callback.message)

@dp.callback_query(F.data.startswith("ashop_toggle_"))
async def admin_toggle_shop(callback: types.CallbackQuery):
    shop_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT is_open FROM shops WHERE id=%s", (shop_id,))
    shop = c.fetchone()
    c.execute("UPDATE shops SET is_open=%s WHERE id=%s", (0 if shop['is_open'] else 1, shop_id))
    conn.commit()
    conn.close()
    await callback.answer("✅ O'zgartirildi")

@dp.callback_query(F.data.startswith("ashop_delete_"))
async def admin_delete_shop(callback: types.CallbackQuery):
    shop_id = int(callback.data.split("_")[2])
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Ha, o'chirish", callback_data=f"ashop_confirm_del_{shop_id}")
    kb.button(text="❌ Yo'q", callback_data=f"admin_shop_{shop_id}")
    await callback.message.answer("⚠️ Rostdan ham do'konni o'chirmoqchimisiz?", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("ashop_confirm_del_"))
async def admin_confirm_delete_shop(callback: types.CallbackQuery):
    shop_id = int(callback.data.split("_")[3])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT owner_tg_id, name FROM shops WHERE id=%s", (shop_id,))
    shop = c.fetchone()
    c.execute("DELETE FROM shops WHERE id=%s", (shop_id,))
    conn.commit()
    conn.close()
    if shop:
        try:
            await bot.send_message(shop['owner_tg_id'], f"⚠️ {shop['name']} do'koningiz tizimdan o'chirildi.")
        except:
            pass
    log_admin_action(callback.from_user.id, "Do'kon o'chirildi", f"Do'kon ID: {shop_id}")
    await callback.message.answer("✅ Do'kon o'chirildi")

@dp.callback_query(F.data.startswith("ashop_percent_"))
async def admin_set_percent(callback: types.CallbackQuery, state: FSMContext):
    shop_id = int(callback.data.split("_")[2])
    await state.update_data(shop_id=shop_id)
    await callback.message.answer("📊 Admin % kiriting (0-100):")
    await state.set_state(AdminPercentState.percent)

@dp.message(AdminPercentState.percent)
async def save_admin_percent(message: types.Message, state: FSMContext):
    try:
        percent = float(message.text)
        if not 0 <= percent <= 100:
            raise ValueError
    except:
        await message.answer("❌ 0 dan 100 gacha raqam kiriting:")
        return

    data = await state.get_data()
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE shops SET admin_percent=%s WHERE id=%s", (percent, data['shop_id']))
    c.execute("SELECT owner_tg_id FROM shops WHERE id=%s", (data['shop_id'],))
    shop = c.fetchone()
    conn.commit()
    conn.close()

    if shop:
        try:
            await bot.send_message(shop['owner_tg_id'], f"📊 Admin ulushi: {percent}% ga o'zgartirildi")
        except:
            pass

    await message.answer(f"✅ Admin % = {percent}% belgilandi")
    await state.clear()

@dp.callback_query(F.data.startswith("ashop_edit_"))
async def admin_edit_shop_field(callback: types.CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    field = parts[2]
    shop_id = int(parts[3])
    await state.update_data(field=field, shop_id=shop_id)
    await callback.message.answer("✍️ Yangi qiymat kiriting:")
    await state.set_state(EditShopAdminState.value)

@dp.message(EditShopAdminState.value)
async def save_admin_shop_edit(message: types.Message, state: FSMContext):
    data = await state.get_data()
    field_map = {"name": "name", "card": "card_number", "phone": "phone"}
    db_field = field_map.get(data['field'], data['field'])
    conn = get_db()
    c = conn.cursor()
    c.execute(f"UPDATE shops SET {db_field}=%s WHERE id=%s", (message.text, data['shop_id']))
    conn.commit()
    conn.close()
    await message.answer("✅ Saqlandi!")
    await state.clear()

@dp.callback_query(F.data.startswith("ashop_orders_"))
async def admin_shop_orders(callback: types.CallbackQuery):
    shop_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM orders WHERE shop_id=%s ORDER BY created_at DESC LIMIT 30", (shop_id,))
    orders = c.fetchall()
    conn.close()

    if not orders:
        await callback.message.answer("📦 Buyurtmalar yo'q")
        return

    kb = InlineKeyboardBuilder()
    for o in orders:
        status_map = {"pending": "⏳", "confirmed": "✅", "on_way": "🚗", "delivered": "✅", "rejected": "❌"}
        emoji = status_map.get(o['status'], "❓")
        kb.button(text=f"{emoji} #{o['order_uid']} {o['created_at']}", callback_data=f"admin_order_{o['order_uid']}")
    kb.button(text="⬅️ Orqaga", callback_data=f"admin_shop_{shop_id}")
    kb.adjust(1)
    await callback.message.answer("📦 Buyurtmalar:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("admin_order_"))
async def admin_order_detail(callback: types.CallbackQuery):
    order_uid = callback.data.split("_", 2)[2]
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM orders WHERE order_uid=%s", (order_uid,))
    o = c.fetchone()
    conn.close()

    if not o:
        await callback.message.answer("Topilmadi")
        return

    user = get_user(o['user_tg_id'])
    status_map = {"pending": "⏳ Kutilmoqda", "confirmed": "✅ Tasdiqlandi", "on_way": "🚗 Yo'lda",
                  "delivered": "✅ Yetkazildi", "rejected": "❌ Rad etildi"}

    text = (f"📦 #{o['order_uid']}\n"
            f"👤 {user['full_name'] if user else 'Tel buyurtma'}\n"
            f"📱 {user['phone'] if user else ''}\n"
            f"📍 {o['address']}\n"
            f"🛍️ {o['products']}\n"
            f"💰 {o['total_sum']:,.0f} so'm\n"
            f"💳 {o['payment_type']}\n"
            f"📅 {o['created_at']}\n"
            f"📊 {status_map.get(o['status'], o['status'])}")

    await callback.message.answer(text)

@dp.callback_query(F.data == "add_new_shop")
async def add_new_shop(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.answer("👤 Do'kon egasining Telegram ID sini kiriting:")
    await state.set_state(AddShopState.tg_id)

@dp.message(AddShopState.tg_id)
async def add_shop_tg(message: types.Message, state: FSMContext):
    try:
        tg_id = int(message.text)
    except:
        await message.answer("❌ Noto'g'ri ID. Raqam kiriting:")
        return
    await state.update_data(owner_tg_id=tg_id)
    await message.answer("🏪 Do'kon nomini kiriting:")
    await state.set_state(AddShopState.name)

@dp.message(AddShopState.name)
async def add_shop_name(message: types.Message, state: FSMContext):
    data = await state.get_data()
    uid = gen_id()
    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO shops (id, owner_tg_id, name, created_at) VALUES (%s,%s,%s,%s)",
              (uid, data['owner_tg_id'], message.text, now_str()))
    conn.commit()
    conn.close()

    conn2 = get_db()
    c2 = conn2.cursor()
    c2.execute("SELECT tg_id FROM users")
    users = c2.fetchall()
    conn2.close()

    for u in users:
        try:
            await bot.send_message(u['tg_id'],
                                   f"🆕 Yangi do'kon ochildi!\n🏪 {message.text}\nTez buyurtma bering!")
        except:
            pass

    try:
        await bot.send_message(data['owner_tg_id'],
                               f"🎉 Do'koningiz tizimga qo'shildi!\n🏪 {message.text}\n🆔 {uid}\n\n"
                               f"Ish boshlash uchun:\n1. Karta raqam\n2. Ish vaqti\n3. Telefon\n4. Mahsulotlar",
                               reply_markup=shop_main_kb())
    except:
        pass

    await message.answer(f"✅ Do'kon qo'shildi! ID: {uid}")
    log_admin_action(message.from_user.id, "Do'kon qo'shdi", f"Do'kon: {message.text}, ID: {uid}")
    await state.clear()

# --- ADMIN: MIJOZLAR ---
@dp.message(F.text == "👥 Mijozlar")
async def admin_users(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM users ORDER BY registered_at DESC")
    users = c.fetchall()
    conn.close()

    kb = InlineKeyboardBuilder()
    for u in users:
        status = "🚫" if u['is_blocked'] else "👤"
        kb.button(text=f"{status} {u['full_name']} ({u['id']})", callback_data=f"auser_{u['tg_id']}")
    kb.adjust(1)
    await message.answer(f"👥 Mijozlar ({len(users)} ta):", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("auser_"))
async def admin_user_detail(callback: types.CallbackQuery):
    tg_id = int(callback.data.split("_")[1])
    user = get_user(tg_id)
    if not user:
        await callback.message.answer("Topilmadi")
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) as cnt FROM orders WHERE user_tg_id=%s", (tg_id,))
    order_count = c.fetchone()['cnt']
    c.execute("SELECT MAX(created_at) as last FROM orders WHERE user_tg_id=%s", (tg_id,))
    last_order = c.fetchone()['last']
    conn.close()

    text = (f"👤 {user['full_name']}\n"
            f"📱 {user['phone']}\n"
            f"🆔 ID: {user['id']}\n"
            f"TG: {tg_id}\n"
            f"@{user['username'] or 'yoq'}\n"
            f"📅 Ro'yxat: {user['registered_at']}\n"
            f"📦 Buyurtmalar: {order_count}\n"
            f"🕐 Oxirgi: {last_order or 'Yo\'q'}\n"
            f"{'🚫 Bloklangan' if user['is_blocked'] else '✅ Faol'}")

    kb = InlineKeyboardBuilder()
    if user['is_blocked']:
        kb.button(text="✅ Blokdan chiqarish", callback_data=f"unblock_user_{tg_id}")
    else:
        kb.button(text="🚫 Bloklash", callback_data=f"block_user_{tg_id}")
    kb.button(text="💬 Chat boshlash", callback_data=f"admin_chat_user_{tg_id}")
    kb.adjust(1)
    await callback.message.answer(text, reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("block_user_"))
async def block_user(callback: types.CallbackQuery):
    tg_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE users SET is_blocked=1 WHERE tg_id=%s", (tg_id,))
    conn.commit()
    conn.close()
    log_admin_action(callback.from_user.id, "Mijoz bloklandi", f"TG ID: {tg_id}")
    try:
        await bot.send_message(tg_id, "🚫 Siz bloklangansiz.")
    except:
        pass
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Blokdan chiqarish", callback_data=f"unblock_user_{tg_id}")
    kb.button(text="💬 Chat boshlash", callback_data=f"admin_chat_user_{tg_id}")
    kb.adjust(1)
    await callback.message.edit_reply_markup(reply_markup=kb.as_markup())
    await callback.answer("🚫 Bloklandi")

@dp.callback_query(F.data.startswith("unblock_user_"))
async def unblock_user(callback: types.CallbackQuery):
    tg_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE users SET is_blocked=0 WHERE tg_id=%s", (tg_id,))
    conn.commit()
    conn.close()
    log_admin_action(callback.from_user.id, "Mijoz blokdan chiqarildi", f"TG ID: {tg_id}")
    try:
        await bot.send_message(tg_id, "✅ Bloklangiz olib tashlandi!")
    except:
        pass
    kb = InlineKeyboardBuilder()
    kb.button(text="🚫 Bloklash", callback_data=f"block_user_{tg_id}")
    kb.button(text="💬 Chat boshlash", callback_data=f"admin_chat_user_{tg_id}")
    kb.adjust(1)
    await callback.message.edit_reply_markup(reply_markup=kb.as_markup())
    await callback.answer("✅ Blokdan chiqarildi")

# --- ADMIN: KURYERLAR ---
@dp.message(F.text == "🚚 Kuryerlar")
async def admin_couriers(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT couriers.*, shops.name as shop_name FROM couriers LEFT JOIN shops ON couriers.shop_id=shops.id")
    couriers = c.fetchall()
    conn.close()

    kb = InlineKeyboardBuilder()
    for cur in couriers:
        status = "🚫" if cur['is_blocked'] else ("🔴" if cur['is_busy'] else "🟢")
        kb.button(text=f"{status} {cur['full_name']} ({cur['shop_name'] or 'Do\'konsiz'})",
                  callback_data=f"acourier_{cur['tg_id']}")
    kb.button(text="➕ Kuryer qo'shish", callback_data="add_courier")
    kb.adjust(1)
    await message.answer(f"🚚 Kuryerlar ({len(couriers)} ta):", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("acourier_"))
async def admin_courier_detail(callback: types.CallbackQuery):
    tg_id = int(callback.data.split("_")[1])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT couriers.*, shops.name as shop_name FROM couriers LEFT JOIN shops ON couriers.shop_id=shops.id WHERE couriers.tg_id=%s", (tg_id,))
    cur = c.fetchone()
    c.execute("SELECT COUNT(*) as cnt FROM orders WHERE courier_tg_id=%s AND status='delivered'", (tg_id,))
    delivered = c.fetchone()['cnt']
    conn.close()

    if not cur:
        await callback.message.answer("Topilmadi")
        return

    text = (f"🚚 {cur['full_name']}\n"
            f"📱 {cur['phone']}\n"
            f"🆔 ID: {cur['id']}\n"
            f"🏪 Do'kon: {cur['shop_name'] or 'Yo\'q'}\n"
            f"📦 Yetkazgan: {delivered}\n"
            f"{'🚫 Bloklangan' if cur['is_blocked'] else '✅ Faol'}\n"
            f"{'🔴 Band' if cur['is_busy'] else '🟢 Bo\'sh'}")

    kb = InlineKeyboardBuilder()
    if cur['is_blocked']:
        kb.button(text="✅ Blokdan chiqarish", callback_data=f"unblock_courier_{tg_id}")
    else:
        kb.button(text="🚫 Bloklash", callback_data=f"block_courier_{tg_id}")
    kb.button(text="🗑️ O'chirish", callback_data=f"delete_courier_{tg_id}")
    kb.adjust(2)
    await callback.message.answer(text, reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("block_courier_"))
async def block_courier(callback: types.CallbackQuery):
    tg_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE couriers SET is_blocked=1 WHERE tg_id=%s", (tg_id,))
    conn.commit()
    conn.close()
    await callback.answer("🚫 Bloklandi")

@dp.callback_query(F.data.startswith("unblock_courier_"))
async def unblock_courier(callback: types.CallbackQuery):
    tg_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE couriers SET is_blocked=0 WHERE tg_id=%s", (tg_id,))
    conn.commit()
    conn.close()
    await callback.answer("✅ Blokdan chiqarildi")

@dp.callback_query(F.data.startswith("delete_courier_"))
async def delete_courier_confirm(callback: types.CallbackQuery):
    tg_id = int(callback.data.split("_")[2])
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Ha, o'chirish", callback_data=f"confirm_del_courier_{tg_id}")
    kb.button(text="❌ Bekor qilish", callback_data="cancel_del_courier")
    kb.adjust(2)
    await callback.message.answer("⚠️ Rostdan ham bu kuryerni o'chirmoqchimisiz?", reply_markup=kb.as_markup())
    await callback.answer()

@dp.callback_query(F.data.startswith("confirm_del_courier_"))
async def delete_courier_execute(callback: types.CallbackQuery):
    tg_id = int(callback.data.split("_")[3])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT full_name FROM couriers WHERE tg_id=%s", (tg_id,))
    cur = c.fetchone()
    c.execute("DELETE FROM couriers WHERE tg_id=%s", (tg_id,))
    conn.commit()
    conn.close()
    try:
        await bot.send_message(tg_id, "❌ Siz kuryerlar ro'yxatidan chiqarildingiz.")
    except:
        pass
    name = cur['full_name'] if cur else str(tg_id)
    await callback.message.edit_text(f"🗑️ {name} o'chirildi.")
    await callback.answer("🗑️ O'chirildi")

@dp.callback_query(F.data == "cancel_del_courier")
async def cancel_del_courier(callback: types.CallbackQuery):
    await callback.message.edit_text("❌ Bekor qilindi.")
    await callback.answer()

@dp.callback_query(F.data == "add_courier")
async def add_courier_start(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.answer("👤 Kuryer ism familyasini kiriting:")
    await state.set_state(AddCourierState.name)

@dp.message(AddCourierState.name)
async def add_courier_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text)
    await message.answer("📱 Telefon raqamini kiriting:")
    await state.set_state(AddCourierState.phone)

@dp.message(AddCourierState.phone)
async def add_courier_phone(message: types.Message, state: FSMContext):
    await state.update_data(phone=message.text)
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, name FROM shops")
    shops = c.fetchall()
    conn.close()

    kb = InlineKeyboardBuilder()
    for s in shops:
        kb.button(text=s['name'], callback_data=f"courier_shop_{s['id']}")
    kb.adjust(1)
    await message.answer("🏪 Qaysi do'konga tayinlansin?", reply_markup=kb.as_markup())
    await state.set_state(AddCourierState.shop_id)

@dp.callback_query(F.data.startswith("courier_shop_"), StateFilter(AddCourierState.shop_id))
async def add_courier_shop(callback: types.CallbackQuery, state: FSMContext):
    shop_id = int(callback.data.split("_")[2])
    data = await state.get_data()
    uid = gen_id()

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) as cnt FROM couriers WHERE shop_id=%s", (shop_id,))
    queue_order = c.fetchone()['cnt']
    c.execute("INSERT INTO couriers (id, tg_id, full_name, phone, shop_id, queue_order, registered_at) VALUES (%s,%s,%s,%s,%s,%s,%s)",
              (uid, uid, data['name'], data['phone'], shop_id, queue_order, now_str()))
    conn.commit()
    conn.close()

    await callback.message.answer(
        f"✅ Kuryer qo'shildi!\n👤 {data['name']}\n📱 {data['phone']}\n\n"
        f"TG ID ni yangilash:\n/set_courier_id {uid} [TG_ID]"
    )
    await state.clear()

@dp.message(Command("set_courier_id"))
async def set_courier_id(message: types.Message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.split()
    if len(parts) != 3:
        await message.answer("Format: /set_courier_id [courier_id] [tg_id]")
        return
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE couriers SET tg_id=%s WHERE id=%s", (int(parts[2]), int(parts[1])))
    conn.commit()
    conn.close()
    await message.answer("✅ Yangilandi!")

# --- ADMIN: STATISTIKA ---
@dp.message(F.text == "📊 Statistika")
async def admin_stats(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM shops")
    shops = c.fetchall()
    c.execute("SELECT COUNT(*) as cnt FROM users")
    users_count = c.fetchone()['cnt']
    c.execute("SELECT COUNT(*) as cnt FROM couriers")
    couriers_count = c.fetchone()['cnt']
    c.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(total_sum),0) as total FROM orders WHERE status='delivered'")
    all_orders = c.fetchone()
    today = datetime.now().strftime("%d.%m.%Y")
    c.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(total_sum),0) as total FROM orders WHERE status='delivered' AND created_at LIKE %s",
              (f"{today}%",))
    today_orders = c.fetchone()
    conn.close()

    text = "📊 Umumiy statistika\n\n"
    for s in shops:
        conn2 = get_db()
        c2 = conn2.cursor()
        c2.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(total_sum),0) as total FROM orders WHERE shop_id=%s AND status='delivered' AND created_at LIKE %s",
                   (s['id'], f"{today}%"))
        sr_today = c2.fetchone()
        c2.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(total_sum),0) as total FROM orders WHERE shop_id=%s AND status='delivered'",
                   (s['id'],))
        sr = c2.fetchone()
        conn2.close()
        text += f"🏪 {s['name']}: bugun {sr_today['cnt']} ta | jami {sr['cnt']} ta ({sr['total']:,.0f} so'm)\n"

    text += f"\n👥 Mijozlar: {users_count}\n"
    text += f"🚚 Kuryerlar: {couriers_count}\n"
    text += f"📦 Jami buyurtmalar: {all_orders['cnt']}\n"
    text += f"💰 Jami daromad: {all_orders['total']:,.0f} so'm\n"
    text += f"📅 Bugungi daromad: {today_orders['total']:,.0f} so'm"

    await message.answer(text)

# --- BROADCAST ---
@dp.message(F.text == "📣 Xabar yuborish")
async def broadcast_menu(message: types.Message):
    if not is_admin(message.from_user.id):
        return
    kb = InlineKeyboardBuilder()
    kb.button(text="👥 Hammaga", callback_data="broadcast_all")
    kb.button(text="🏪 Do'kon egalariga", callback_data="broadcast_shops")
    kb.button(text="🚚 Kuryerlarga", callback_data="broadcast_couriers")
    kb.adjust(1)
    await message.answer("📣 Kimga xabar?", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("broadcast_"))
async def broadcast_target(callback: types.CallbackQuery, state: FSMContext):
    target = callback.data.split("_")[1]
    await state.update_data(target=target)
    await callback.message.answer("✍️ Xabarni yozing:")
    await state.set_state(BroadcastState.message)

@dp.message(BroadcastState.message)
async def broadcast_send(message: types.Message, state: FSMContext):
    data = await state.get_data()
    target = data['target']

    conn = get_db()
    c = conn.cursor()
    ids = []
    if target == "all":
        c.execute("SELECT tg_id FROM users")
        ids = [r['tg_id'] for r in c.fetchall()]
        c.execute("SELECT owner_tg_id FROM shops")
        ids += [r['owner_tg_id'] for r in c.fetchall()]
        c.execute("SELECT tg_id FROM couriers")
        ids += [r['tg_id'] for r in c.fetchall()]
    elif target == "shops":
        c.execute("SELECT owner_tg_id FROM shops")
        ids = [r['owner_tg_id'] for r in c.fetchall()]
    elif target == "couriers":
        c.execute("SELECT tg_id FROM couriers")
        ids = [r['tg_id'] for r in c.fetchall()]
    conn.close()

    sent = 0
    for uid in set(ids):
        try:
            if message.photo:
                await bot.send_photo(uid, message.photo[-1].file_id, caption=message.caption or "")
            else:
                await bot.send_message(uid, message.text)
            sent += 1
        except:
            pass

    await message.answer(f"✅ {sent} ta foydalanuvchiga yuborildi")
    await state.clear()

# --- SEARCH ---
@dp.message(F.text == "🔍 Qidirish")
async def search_menu(message: types.Message):
    if not is_admin(message.from_user.id):
        return
    kb = InlineKeyboardBuilder()
    kb.button(text="👥 Mijozlar", callback_data="search_users")
    kb.button(text="🏪 Do'konlar", callback_data="search_shops")
    kb.button(text="🚚 Kuryerlar", callback_data="search_couriers")
    kb.adjust(3)
    await message.answer("🔍 Qidiruv turi:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("search_"))
async def search_start(callback: types.CallbackQuery, state: FSMContext):
    stype = callback.data.split("_")[1]
    await state.update_data(search_type=stype)
    await callback.message.answer("🔍 Qidirish uchun matn kiriting:")
    await state.set_state(SearchState.query)

@dp.message(SearchState.query)
async def do_search(message: types.Message, state: FSMContext):
    data = await state.get_data()
    query = message.text.strip()
    stype = data['search_type']

    conn = get_db()
    c = conn.cursor()

    if stype == "users":
        c.execute("SELECT * FROM users WHERE full_name ILIKE %s OR phone ILIKE %s OR CAST(id AS TEXT) ILIKE %s OR CAST(tg_id AS TEXT) ILIKE %s",
                  (f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%"))
        results = c.fetchall()
        text = "👥 Topilganlar:\n\n" + "\n".join([f"👤 {u['full_name']} | {u['phone']} | {u['id']}" for u in results]) if results else "❌ Topilmadi"
    elif stype == "shops":
        c.execute("SELECT * FROM shops WHERE name ILIKE %s OR CAST(id AS TEXT) ILIKE %s", (f"%{query}%", f"%{query}%"))
        results = c.fetchall()
        text = "🏪 Topilganlar:\n\n" + "\n".join([f"🏪 {s['name']} | {s['id']}" for s in results]) if results else "❌ Topilmadi"
    elif stype == "couriers":
        c.execute("SELECT * FROM couriers WHERE full_name ILIKE %s OR phone ILIKE %s OR CAST(tg_id AS TEXT) ILIKE %s",
                  (f"%{query}%", f"%{query}%", f"%{query}%"))
        results = c.fetchall()
        text = "🚚 Topilganlar:\n\n" + "\n".join([f"🚚 {r['full_name']} | {r['phone']}" for r in results]) if results else "❌ Topilmadi"
    else:
        text = "❌ Noto'g'ri tur"

    conn.close()
    await message.answer(text)
    await state.clear()

# --- ORDERS ---
@dp.message(F.text == "📋 Buyurtmalar")
async def admin_orders(message: types.Message):
    if not is_admin(message.from_user.id):
        return
    kb = InlineKeyboardBuilder()
    kb.button(text="⏳ Kutilmoqda", callback_data="orders_status_pending")
    kb.button(text="🚗 Yo'lda", callback_data="orders_status_on_way")
    kb.button(text="✅ Yetkazildi", callback_data="orders_status_delivered")
    kb.adjust(3)
    await message.answer("📋 Buyurtmalar holati:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("orders_status_"))
async def orders_by_status(callback: types.CallbackQuery):
    status = callback.data.split("_", 2)[2]
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM orders WHERE status=%s ORDER BY created_at DESC LIMIT 30", (status,))
    orders = c.fetchall()
    conn.close()

    if not orders:
        await callback.message.answer("📦 Buyurtmalar yo'q")
        return

    kb = InlineKeyboardBuilder()
    for o in orders:
        kb.button(text=f"#{o['order_uid']} — {o['created_at']}", callback_data=f"admin_order_{o['order_uid']}")
    kb.adjust(1)
    await callback.message.answer(f"📋 {len(orders)} ta buyurtma:", reply_markup=kb.as_markup())

# --- MUAMMOLI ---
@dp.message(F.text == "⚠️ Muammoli")
async def problematic_orders(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT orders.*, shops.name as shop_name FROM orders LEFT JOIN shops ON orders.shop_id=shops.id WHERE orders.status='pending' ORDER BY orders.created_at ASC LIMIT 20")
    pending = c.fetchall()
    c.execute("SELECT COUNT(*) as cnt FROM orders WHERE status='rejected'")
    rejected_cnt = c.fetchone()['cnt']
    conn.close()

    text = f"⚠️ Muammoli buyurtmalar\n\n⏳ Kutilayotgan: {len(pending)} ta\n❌ Rad etilgan: {rejected_cnt} ta\n\n"

    for o in pending[:10]:
        try:
            created = datetime.strptime(o['created_at'], "%d.%m.%Y %H:%M")
            wait_min = int((datetime.now() - created).total_seconds() / 60)
            warn = " 🔴" if wait_min >= 10 else ""
            text += f"  #{o['order_uid']} | {o['shop_name'] or '?'} | {wait_min} daqiqa{warn}\n"
        except:
            text += f"  #{o['order_uid']} | {o['created_at']}\n"

    await message.answer(text)

# --- MOLIYA ---
@dp.message(F.text == "💰 Moliya")
async def finance_menu(message: types.Message):
    if not is_admin(message.from_user.id):
        return
    kb = InlineKeyboardBuilder()
    kb.button(text="💳 Karta to'lovlari", callback_data="finance_card")
    kb.button(text="💵 Naqd to'lovlar", callback_data="finance_cash")
    kb.adjust(2)
    await message.answer("💰 Moliya:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("finance_"))
async def finance_detail(callback: types.CallbackQuery):
    pay_type = "Karta" if callback.data == "finance_card" else "Naqd"
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT orders.*, shops.name as shop_name FROM orders LEFT JOIN shops ON orders.shop_id=shops.id WHERE payment_type=%s AND status='delivered' ORDER BY created_at DESC LIMIT 20", (pay_type,))
    orders = c.fetchall()
    c.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(total_sum),0) as total FROM orders WHERE payment_type=%s AND status='delivered'", (pay_type,))
    summary = c.fetchone()
    conn.close()

    text = f"{'💳' if pay_type == 'Karta' else '💵'} {pay_type} to'lovlar\n\n"
    text += f"📦 Jami: {summary['cnt']} ta\n💰 Summa: {summary['total']:,.0f} so'm\n\n"
    for o in orders[:10]:
        text += f"#{o['order_uid']} | {o['shop_name']} | {o['total_sum']:,.0f} so'm\n"

    await callback.message.answer(text)

# --- EXCEL EKSPORT (TUZATILGAN) ---
@dp.message(F.text == "📥 Excel eksport")
async def excel_export(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    log_admin_action(message.from_user.id, "Excel eksport")
    await message.answer("⏳ Excel fayl tayyorlanmoqda...")

    try:
        # Barcha ma'lumotlarni BIR MARTA olish
        conn = get_db()
        c = conn.cursor()

        c.execute("""
            SELECT o.*, s.name as shop_name,
                   cu.full_name as courier_name,
                   u.full_name as user_name, u.phone as user_phone
            FROM orders o
            LEFT JOIN shops s ON o.shop_id = s.id
            LEFT JOIN couriers cu ON o.courier_tg_id = cu.tg_id
            LEFT JOIN users u ON o.user_tg_id = u.tg_id
            ORDER BY o.created_at DESC
        """)
        orders = c.fetchall()

        c.execute("""
            SELECT u.*,
                   COUNT(o.id) as order_count,
                   COALESCE(SUM(o.total_sum), 0) as total_spent
            FROM users u
            LEFT JOIN orders o ON u.tg_id = o.user_tg_id AND o.status = 'delivered'
            GROUP BY u.id, u.tg_id, u.username, u.full_name, u.phone, u.registered_at, u.is_blocked
            ORDER BY order_count DESC
        """)
        users = c.fetchall()

        c.execute("""
            SELECT cu.*, s.name as shop_name,
                   COUNT(o.id) as delivered_count,
                   COALESCE(SUM(o.total_sum), 0) as total_sum
            FROM couriers cu
            LEFT JOIN shops s ON cu.shop_id = s.id
            LEFT JOIN orders o ON cu.tg_id = o.courier_tg_id AND o.status = 'delivered'
            GROUP BY cu.id, cu.tg_id, cu.full_name, cu.phone, cu.shop_id,
                     cu.is_busy, cu.is_blocked, cu.is_available, cu.queue_order, cu.registered_at, s.name
            ORDER BY delivered_count DESC
        """)
        couriers = c.fetchall()

        c.execute("""
            SELECT s.*,
                   COUNT(o.id) as order_count,
                   COALESCE(SUM(CASE WHEN o.status='delivered' THEN o.total_sum ELSE 0 END), 0) as total_income
            FROM shops s
            LEFT JOIN orders o ON s.id = o.shop_id
            GROUP BY s.id, s.owner_tg_id, s.name, s.phone, s.card_number, s.work_time,
                     s.is_open, s.rating, s.rating_count, s.admin_percent, s.created_at, s.vacation_until
            ORDER BY total_income DESC
        """)
        shops = c.fetchall()

        c.execute("SELECT * FROM promo_codes ORDER BY created_at DESC")
        promos = c.fetchall()

        c.execute("SELECT * FROM admin_logs ORDER BY created_at DESC LIMIT 500")
        admin_logs = c.fetchall()

        c.execute("SELECT * FROM chats ORDER BY created_at DESC LIMIT 500")
        chats = c.fetchall()

        conn.close()

        # Excel yaratish
        wb = openpyxl.Workbook()

        h_font = Font(bold=True, color="FFFFFF", size=11)
        h_fill = PatternFill("solid", fgColor="1A3C5E")
        center = Alignment(horizontal="center", vertical="center")

        def make_header(ws, headers, widths=None):
            for i, h in enumerate(headers, 1):
                cell = ws.cell(1, i, h)
                cell.font = h_font
                cell.fill = h_fill
                cell.alignment = center
                ws.column_dimensions[cell.column_letter].width = widths[i-1] if widths and i <= len(widths) else max(len(str(h))+4, 14)

        status_map = {
            "pending": "Kutilmoqda", "confirmed": "Tasdiqlandi",
            "on_way": "Yo'lda", "delivered": "Yetkazildi", "rejected": "Rad etildi"
        }

        # 1. DO'KONLAR
        ws1 = wb.active
        ws1.title = "Do'konlar"
        make_header(ws1, ["ID", "Nom", "Egasi TG", "Tel", "Karta", "Ish vaqti",
                           "Reyting", "Buyurtmalar", "Daromad", "Admin %", "Holat", "Yaratilgan"],
                    [10, 22, 14, 14, 22, 14, 10, 14, 16, 10, 12, 16])
        for i, s in enumerate(shops, 2):
            ws1.cell(i, 1, s['id'])
            ws1.cell(i, 2, s['name'])
            ws1.cell(i, 3, s['owner_tg_id'])
            ws1.cell(i, 4, s['phone'] or '')
            ws1.cell(i, 5, s['card_number'] or '')
            ws1.cell(i, 6, s['work_time'] or '')
            ws1.cell(i, 7, round(s['rating'], 2))
            ws1.cell(i, 8, s['order_count'])
            ws1.cell(i, 9, s['total_income'])
            ws1.cell(i, 10, s['admin_percent'])
            ws1.cell(i, 11, 'Ochiq' if s['is_open'] else 'Yopiq')
            ws1.cell(i, 12, s['created_at'])

        # 2. BUYURTMALAR
        ws2 = wb.create_sheet("Buyurtmalar")
        make_header(ws2, ["ID", "Do'kon", "Mijoz", "Tel", "Mahsulotlar",
                           "Summa", "Chegirma", "To'lov", "Manzil", "Holat",
                           "Kuryer", "Buyurtma vaqti", "Yetkazildi", "Manba"],
                    [10, 18, 16, 14, 30, 14, 12, 10, 20, 14, 16, 16, 16, 12])
        for i, o in enumerate(orders, 2):
            ws2.cell(i, 1, o['order_uid'])
            ws2.cell(i, 2, o['shop_name'] or '')
            ws2.cell(i, 3, o['user_name'] or ('Tel buyurtma' if o['user_tg_id'] == 0 else str(o['user_tg_id'])))
            ws2.cell(i, 4, o['user_phone'] or '')
            ws2.cell(i, 5, o['products'])
            ws2.cell(i, 6, o['total_sum'])
            ws2.cell(i, 7, o['discount'] or 0)
            ws2.cell(i, 8, o['payment_type'])
            ws2.cell(i, 9, o['address'])
            ws2.cell(i, 10, status_map.get(o['status'], o['status']))
            ws2.cell(i, 11, o['courier_name'] or '')
            ws2.cell(i, 12, o['created_at'])
            ws2.cell(i, 13, o['delivered_at'] or '')
            ws2.cell(i, 14, o['source'] or 'normal')

        # 3. KURYERLAR
        ws3 = wb.create_sheet("Kuryerlar")
        make_header(ws3, ["ID", "Ism", "Tel", "TG ID", "Do'kon",
                           "Yetkazgan", "Jami summa", "Holat", "Qo'shilgan"],
                    [10, 20, 14, 13, 18, 12, 16, 12, 16])
        for i, cur in enumerate(couriers, 2):
            ws3.cell(i, 1, cur['id'])
            ws3.cell(i, 2, cur['full_name'])
            ws3.cell(i, 3, cur['phone'])
            ws3.cell(i, 4, cur['tg_id'])
            ws3.cell(i, 5, cur['shop_name'] or '')
            ws3.cell(i, 6, cur['delivered_count'])
            ws3.cell(i, 7, cur['total_sum'])
            ws3.cell(i, 8, 'Bloklangan' if cur['is_blocked'] else 'Faol')
            ws3.cell(i, 9, cur['registered_at'] or '')

        # 4. MIJOZLAR
        ws4 = wb.create_sheet("Mijozlar")
        make_header(ws4, ["ID", "Ism", "Tel", "Username", "TG ID",
                           "Buyurtmalar", "Jami xarid", "Ro'yxat", "Holat"],
                    [10, 20, 14, 14, 13, 14, 16, 16, 12])
        for i, u in enumerate(users, 2):
            ws4.cell(i, 1, u['id'])
            ws4.cell(i, 2, u['full_name'])
            ws4.cell(i, 3, u['phone'])
            ws4.cell(i, 4, u['username'] or '')
            ws4.cell(i, 5, u['tg_id'])
            ws4.cell(i, 6, u['order_count'])
            ws4.cell(i, 7, u['total_spent'])
            ws4.cell(i, 8, u['registered_at'])
            ws4.cell(i, 9, 'Bloklangan' if u['is_blocked'] else 'Faol')

        # 5. PROMO KODLAR
        ws5 = wb.create_sheet("Promo kodlar")
        make_header(ws5, ["Kod", "Chegirma", "Tur", "Min summa",
                           "Muddat", "Limit", "Ishlatilgan", "Yaratilgan"],
                    [14, 12, 12, 14, 14, 10, 14, 16])
        for i, p in enumerate(promos, 2):
            ws5.cell(i, 1, p['code'])
            ws5.cell(i, 2, p['discount_value'])
            ws5.cell(i, 3, p['discount_type'])
            ws5.cell(i, 4, p['min_sum'] or 0)
            ws5.cell(i, 5, p['expires_at'] or 'Cheksiz')
            ws5.cell(i, 6, p['max_uses'] or 0)
            ws5.cell(i, 7, p['used_count'] or 0)
            ws5.cell(i, 8, p['created_at'] or '')

        # 6. ADMIN LOGLARI
        ws6 = wb.create_sheet("Admin loglari")
        make_header(ws6, ["Admin TG", "Harakat", "Tafsilot", "Vaqt"],
                    [13, 25, 40, 16])
        for i, lg in enumerate(admin_logs, 2):
            ws6.cell(i, 1, lg['admin_tg_id'])
            ws6.cell(i, 2, lg['action'])
            ws6.cell(i, 3, lg['details'] or '')
            ws6.cell(i, 4, lg['created_at'])

        # 7. MOLIYA XULOSASI
        ws7 = wb.create_sheet("Moliya")
        make_header(ws7, ["Do'kon", "Buyurtmalar", "Daromad", "Admin %", "Admin ulushi"],
                    [22, 14, 16, 12, 16])
        for i, s in enumerate(shops, 2):
            admin_share = s['total_income'] * s['admin_percent'] / 100
            ws7.cell(i, 1, s['name'])
            ws7.cell(i, 2, s['order_count'])
            ws7.cell(i, 3, s['total_income'])
            ws7.cell(i, 4, s['admin_percent'])
            ws7.cell(i, 5, round(admin_share, 0))

        # 8. STATISTIKA
        ws8 = wb.create_sheet("Statistika")
        ws8.column_dimensions['A'].width = 32
        ws8.column_dimensions['B'].width = 20
        sf = Font(bold=True, size=11)

        delivered_orders = [o for o in orders if o['status'] == 'delivered']
        today_str = datetime.now().strftime("%d.%m.%Y")
        today_del = [o for o in delivered_orders if o['created_at'] and o['created_at'].startswith(today_str)]

        stats = [
            ("BOT STATISTIKASI", ""),
            ("", ""),
            ("Jami mijozlar", len(users)),
            ("Jami do'konlar", len(shops)),
            ("Jami kuryerlar", len(couriers)),
            ("Jami buyurtmalar", len(orders)),
            ("Yetkazilgan", len(delivered_orders)),
            ("Bugungi buyurtmalar", len(today_del)),
            ("Jami daromad (so'm)", f"{sum(o['total_sum'] for o in delivered_orders):,.0f}"),
            ("Bugungi daromad (so'm)", f"{sum(o['total_sum'] for o in today_del):,.0f}"),
            ("", ""),
            ("Eksport vaqti", now_str()),
        ]
        for ri, (k, v) in enumerate(stats, 1):
            ws8.cell(ri, 1, k).font = sf
            ws8.cell(ri, 2, v).font = sf

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"olimbek_savdo_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
        await message.answer_document(
            types.BufferedInputFile(buf.read(), filename=filename),
            caption=(f"📥 Excel eksport tayyor!\n\n"
                     f"8 ta varaq:\n"
                     f"1. Do'konlar\n"
                     f"2. Buyurtmalar\n"
                     f"3. Kuryerlar\n"
                     f"4. Mijozlar\n"
                     f"5. Promo kodlar\n"
                     f"6. Admin loglari\n"
                     f"7. Moliya xulosasi\n"
                     f"8. Statistika\n\n"
                     f"📅 {now_str()}")
        )

    except Exception as e:
        logger.error(f"Excel eksport xatosi: {e}", exc_info=True)
        await message.answer(f"❌ Excel tayyorlashda xatolik:\n{e}\n\nAdmin loglarni tekshiring.")

# --- PROMO KODLAR ---
@dp.message(F.text == "🎟️ Promo kodlar")
async def promo_menu(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM promo_codes ORDER BY created_at DESC")
    promos = c.fetchall()
    conn.close()

    kb = InlineKeyboardBuilder()
    for p in promos:
        status = "✅" if (p['max_uses'] == 0 or p['used_count'] < p['max_uses']) else "❌"
        val_text = f"{p['discount_value']}%" if p['discount_type'] == 'percent' else f"{p['discount_value']:,.0f} so'm"
        kb.button(text=f"{status} {p['code']} — {val_text}", callback_data=f"promo_{p['id']}")
    kb.button(text="➕ Promo kod qo'shish", callback_data="add_promo")
    kb.adjust(1)
    await message.answer("🎟️ Promo kodlar:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("promo_") & ~F.data.startswith("promo_del") & ~F.data.startswith("promo_type"))
async def promo_detail(callback: types.CallbackQuery):
    promo_id = int(callback.data.split("_")[1])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM promo_codes WHERE id=%s", (promo_id,))
    p = c.fetchone()
    conn.close()

    if not p:
        await callback.message.answer("Topilmadi")
        return

    text = (f"🎟️ Kod: {p['code']}\n"
            f"💰 Chegirma: {p['discount_value']}{'%' if p['discount_type'] == 'percent' else ' so\'m'}\n"
            f"💵 Min summa: {p['min_sum']:,.0f} so'm\n"
            f"📅 Muddat: {p['expires_at'] or 'Cheksiz'}\n"
            f"👥 Limit: {p['max_uses'] if p['max_uses'] > 0 else 'Cheksiz'}\n"
            f"✅ Ishlatilgan: {p['used_count']}")

    kb = InlineKeyboardBuilder()
    kb.button(text="🗑️ O'chirish", callback_data=f"promo_del_{promo_id}")
    await callback.message.answer(text, reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("promo_del_"))
async def delete_promo(callback: types.CallbackQuery):
    promo_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM promo_codes WHERE id=%s", (promo_id,))
    conn.commit()
    conn.close()
    await callback.message.answer("🗑️ O'chirildi")

@dp.callback_query(F.data == "add_promo")
async def add_promo_start(callback: types.CallbackQuery, state: FSMContext):
    await callback.message.answer("✍️ Promo kod nomini kiriting (masalan: BAHOR10):")
    await state.set_state(PromoState.discount_type)

@dp.message(PromoState.discount_type)
async def promo_code_name(message: types.Message, state: FSMContext):
    await state.update_data(code=message.text.upper())
    kb = InlineKeyboardBuilder()
    kb.button(text="% Foiz", callback_data="promo_type_percent")
    kb.button(text="💵 Summa", callback_data="promo_type_sum")
    await message.answer("💰 Chegirma turi:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("promo_type_"))
async def promo_type(callback: types.CallbackQuery, state: FSMContext):
    dtype = callback.data.split("_")[2]
    await state.update_data(discount_type=dtype)
    await callback.message.answer("💰 Chegirma miqdorini kiriting:")
    await state.set_state(PromoState.discount_value)

@dp.message(PromoState.discount_value)
async def promo_value(message: types.Message, state: FSMContext):
    try:
        val = float(message.text)
    except:
        await message.answer("❌ Raqam kiriting:")
        return
    await state.update_data(discount_value=val)
    await message.answer("💵 Minimal buyurtma summasini kiriting (so'm):")
    await state.set_state(PromoState.min_sum)

@dp.message(PromoState.min_sum)
async def promo_min_sum(message: types.Message, state: FSMContext):
    try:
        val = float(message.text)
    except:
        await message.answer("❌ Raqam kiriting:")
        return
    await state.update_data(min_sum=val)
    await message.answer("📅 Nech kun ishlashi? (0 = cheksiz):")
    await state.set_state(PromoState.days)

@dp.message(PromoState.days)
async def promo_days(message: types.Message, state: FSMContext):
    try:
        days = int(message.text)
    except:
        await message.answer("❌ Raqam kiriting:")
        return
    await state.update_data(days=days)
    await message.answer("👥 Necha kishi ishlatishi? (0 = cheksiz):")
    await state.set_state(PromoState.max_uses)

@dp.message(PromoState.max_uses)
async def promo_max_uses(message: types.Message, state: FSMContext):
    try:
        max_uses = int(message.text)
    except:
        await message.answer("❌ Raqam kiriting:")
        return

    data = await state.get_data()
    expires_at = None
    if data['days'] > 0:
        exp = datetime.now() + timedelta(days=data['days'])
        expires_at = exp.strftime("%d.%m.%Y")

    uid = gen_id()
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO promo_codes (id, code, discount_type, discount_value, min_sum, days, max_uses, created_at, expires_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                  (uid, data['code'], data['discount_type'], data['discount_value'],
                   data['min_sum'], data['days'], max_uses, now_str(), expires_at))
        conn.commit()
        await message.answer(f"✅ Promo kod yaratildi!\n🎟️ {data['code']}\n💰 {data['discount_value']}{'%' if data['discount_type'] == 'percent' else ' so\'m'}")
    except:
        await message.answer("❌ Bu kod allaqachon mavjud!")
    conn.close()
    await state.clear()

# --- MONITORING ---
@dp.message(F.text == "👁️ Monitoring")
async def live_monitoring(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) as cnt FROM orders WHERE status='pending'")
    pending = c.fetchone()['cnt']
    c.execute("SELECT COUNT(*) as cnt FROM orders WHERE status='on_way'")
    on_way = c.fetchone()['cnt']
    c.execute("SELECT COUNT(*) as cnt FROM couriers WHERE is_busy=1 AND is_blocked=0")
    busy = c.fetchone()['cnt']
    c.execute("SELECT COUNT(*) as cnt FROM couriers WHERE is_busy=0 AND is_blocked=0 AND is_available=1")
    free = c.fetchone()['cnt']
    c.execute("SELECT COUNT(*) as cnt FROM shops WHERE is_open=1")
    open_shops = c.fetchone()['cnt']
    one_hour_ago = (datetime.now() - timedelta(hours=1)).strftime("%d.%m.%Y %H:%M")
    c.execute("SELECT COUNT(*) as cnt FROM orders WHERE created_at >= %s", (one_hour_ago,))
    last_hour = c.fetchone()['cnt']
    conn.close()

    text = (f"👁️ Jonli monitoring\n\n"
            f"📦 Faol:\n  ⏳ Kutilmoqda: {pending}\n  🚗 Yo'lda: {on_way}\n\n"
            f"🚚 Kuryerlar:\n  🔴 Band: {busy}\n  🟢 Bo'sh: {free}\n\n"
            f"🏪 Ochiq do'konlar: {open_shops}\n"
            f"🕐 Oxirgi 1 soatda: {last_hour} buyurtma")

    await message.answer(text)

# --- CHATLAR ---
@dp.message(F.text == "💬 Chatlar")
async def admin_chats(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT DISTINCT from_tg_id, to_tg_id, chat_type, MAX(created_at) as last FROM chats GROUP BY from_tg_id, to_tg_id, chat_type ORDER BY last DESC LIMIT 20")
    chats = c.fetchall()
    conn.close()

    if not chats:
        await message.answer("💬 Chatlar yo'q")
        return

    kb = InlineKeyboardBuilder()
    for ch in chats:
        kb.button(text=f"💬 {ch['from_tg_id']} ({ch['chat_type']}) {ch['last']}",
                  callback_data=f"view_chat_{ch['from_tg_id']}_{ch['to_tg_id']}")
    kb.adjust(1)
    await message.answer("💬 Chatlar:", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("view_chat_"))
async def view_chat_detail(callback: types.CallbackQuery):
    parts = callback.data.split("_")
    from_id = int(parts[2])
    to_id = int(parts[3])

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM chats WHERE (from_tg_id=%s AND to_tg_id=%s) OR (from_tg_id=%s AND to_tg_id=%s) ORDER BY created_at ASC LIMIT 30",
              (from_id, to_id, to_id, from_id))
    messages = c.fetchall()
    conn.close()

    text = "💬 Chat:\n\n"
    for m in messages:
        text += f"[{m['created_at']}] {m['from_tg_id']}: {m['message']}\n"

    await callback.message.answer(text[:4000])

# --- BLOKLANGAN ---
@dp.message(F.text == "🚫 Bloklangan")
async def blocked_list(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE is_blocked=1")
    blocked_users = c.fetchall()
    c.execute("SELECT * FROM couriers WHERE is_blocked=1")
    blocked_couriers = c.fetchall()
    conn.close()

    text = f"🚫 Bloklangan\n\n👥 Mijozlar ({len(blocked_users)} ta):\n"
    for u in blocked_users:
        text += f"  👤 {u['full_name']} | {u['id']}\n"
    text += f"\n🚚 Kuryerlar ({len(blocked_couriers)} ta):\n"
    for cur in blocked_couriers:
        text += f"  🚚 {cur['full_name']} | {cur['id']}\n"

    await message.answer(text if len(text) > 50 else "🚫 Hech kim bloklanmagan")

# --- HAFTALIK HISOBOT ---
@dp.message(F.text == "📈 Haftalik hisobot")
async def weekly_report(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    week_ago = (datetime.now() - timedelta(days=7)).strftime("%d.%m.%Y")
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(total_sum),0) as total FROM orders WHERE status='delivered' AND created_at >= %s", (week_ago,))
    result = c.fetchone()
    c.execute("SELECT COUNT(*) as cnt FROM users WHERE registered_at >= %s", (week_ago,))
    new_users = c.fetchone()['cnt']
    conn.close()

    await message.answer(f"📈 Haftalik hisobot\n\n"
                         f"📦 Buyurtmalar: {result['cnt']}\n"
                         f"💰 Daromad: {result['total']:,.0f} so'm\n"
                         f"👥 Yangi mijozlar: {new_users}")

# --- ADMIN LOGLARI ---
@dp.message(F.text == "🔐 Admin loglari")
async def admin_logs_view(message: types.Message):
    if not is_admin(message.from_user.id):
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM admin_logs ORDER BY created_at DESC LIMIT 30")
    logs = c.fetchall()
    conn.close()

    if not logs:
        await message.answer("🔐 Hozircha loglar yo'q")
        return

    text = "🔐 Admin harakatlari:\n\n"
    for lg in logs:
        text += f"👤 {lg['admin_tg_id']} | {lg['action']}"
        if lg['details']:
            text += f" | {lg['details']}"
        text += f"\n🕐 {lg['created_at']}\n\n"

    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        await message.answer(chunk)

# --- BOT MA'LUMOTLARINI O'CHIRISH ---
@dp.message(F.text == "🗑️ Bot ma'lumotlari")
async def delete_bot_data_start(message: types.Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await message.answer("🔐 Parolni kiriting:")
    await state.set_state(DeleteBotState.password)

@dp.message(DeleteBotState.password)
async def delete_bot_check_pass(message: types.Message, state: FSMContext):
    if message.text != "ROZIMAN100":
        await message.answer("❌ Noto'g'ri parol!")
        await state.clear()
        return

    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Tasdiqlash — HAMMA MA'LUMOT O'CHADI", callback_data="confirm_delete_all")
    kb.button(text="❌ Bekor qilish", callback_data="cancel_delete_all")
    kb.adjust(1)
    await message.answer("⚠️ DIQQAT! Bu amalni ortga qaytarib bo'lmaydi!", reply_markup=kb.as_markup())
    await state.clear()

@dp.callback_query(F.data == "confirm_delete_all")
async def confirm_delete_all(callback: types.CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    conn = get_db()
    c = conn.cursor()
    for table in ['users', 'orders', 'products', 'couriers', 'shops',
                  'chats', 'promo_codes', 'monthly_reports', 'carts', 'active_sessions', 'admins']:
        c.execute(f"DELETE FROM {table}")
    conn.commit()
    conn.close()
    await callback.message.answer("✅ Hamma ma'lumotlar o'chirildi!")

@dp.callback_query(F.data == "cancel_delete_all")
async def cancel_delete_all(callback: types.CallbackQuery):
    await callback.message.answer("❌ Bekor qilindi")

# --- PHONE ORDER CALLBACK ---
@dp.callback_query(F.data.startswith("phone_order_"))
async def phone_order_for_shop(callback: types.CallbackQuery):
    shop_id = int(callback.data.split("_")[2])
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT phone FROM shops WHERE id=%s", (shop_id,))
    shop = c.fetchone()
    conn.close()

    if shop and shop['phone']:
        await callback.message.answer(f"📞 Do'kon telefoni: {shop['phone']}")
    else:
        await callback.message.answer("📞 Do'kon telefon raqami ko'rsatilmagan")

# --- ADMIN → MIJOZ CHAT ---
@dp.callback_query(F.data.startswith("admin_chat_user_"))
async def admin_start_chat_with_user(callback: types.CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    user_tg_id = int(callback.data.split("_")[3])
    user = get_user(user_tg_id)
    await state.update_data(user_tg_id=user_tg_id)
    await state.set_state(AdminUserChatState.chatting)

    kb = InlineKeyboardBuilder()
    kb.button(text="🔚 Chatni tugatish", callback_data=f"admin_end_chat_{user_tg_id}")
    await callback.message.answer(
        f"💬 {user['full_name'] if user else user_tg_id} bilan chat boshlandi.\nXabaringizni yozing:",
        reply_markup=kb.as_markup()
    )

@dp.message(AdminUserChatState.chatting)
async def admin_send_to_user(message: types.Message, state: FSMContext):
    data = await state.get_data()
    user_tg_id = data.get('user_tg_id')
    if not user_tg_id:
        await state.clear()
        return

    try:
        kb = InlineKeyboardBuilder()
        kb.button(text="💬 Admin bilan yozishish", callback_data="reply_to_admin_direct")
        await bot.send_message(user_tg_id, f"📩 Admin xabari:\n\n{message.text}", reply_markup=kb.as_markup())
        await message.answer("✅ Xabar yetkazildi")
    except:
        await message.answer("❌ Xabar yuborib bo'lmadi")

    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO chats (from_tg_id, to_tg_id, message, chat_type, created_at) VALUES (%s,%s,%s,%s,%s)",
              (message.from_user.id, user_tg_id, message.text, "admin_to_user", now_str()))
    conn.commit()
    conn.close()

@dp.callback_query(F.data.startswith("admin_end_chat_"))
async def admin_end_chat(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.answer("✅ Chat tugatildi.", reply_markup=admin_main_kb())

@dp.callback_query(F.data == "reply_to_admin_direct")
async def reply_to_admin_direct(callback: types.CallbackQuery, state: FSMContext):
    tg_id = callback.from_user.id
    admin_id = ADMIN_IDS[0]
    chat_set(tg_id, admin_id, "user_admin", "")
    chat_set(admin_id, tg_id, "admin_user", "")
    await state.set_state(ChatState.chatting)
    await callback.message.answer("💬 Admin bilan chat boshlandi. Yozing:", reply_markup=chat_end_kb())
    user = get_user(tg_id)
    await notify_partner_chat_started(admin_id, tg_id, f"Mijoz {user['full_name'] if user else tg_id}", "admin_user")
    await callback.answer()

# --- BACKGROUND TASKS ---
async def check_stuck_orders():
    while True:
        try:
            conn = get_db()
            c = conn.cursor()
            c.execute("SELECT orders.*, shops.name as shop_name FROM orders LEFT JOIN shops ON orders.shop_id=shops.id WHERE orders.status='pending'")
            pending = c.fetchall()
            conn.close()

            for o in pending:
                try:
                    created = datetime.strptime(o['created_at'], "%d.%m.%Y %H:%M")
                    wait_min = int((datetime.now() - created).total_seconds() / 60)
                    if wait_min == 10:
                        alert = (f"🔴 10 daqiqa javob yo'q!\n\n"
                                 f"📦 #{o['order_uid']}\n"
                                 f"🏪 {o['shop_name'] or 'Noma\'lum'}\n"
                                 f"💰 {o['total_sum']:,.0f} so'm\n"
                                 f"📅 {o['created_at']}")
                        for admin_id in ADMIN_IDS:
                            try:
                                await bot.send_message(admin_id, alert)
                            except:
                                pass
                except:
                    pass
        except Exception as e:
            logger.error(f"check_stuck_orders: {e}")
        await asyncio.sleep(60)

async def check_user_order_timeout():
    while True:
        try:
            conn = get_db()
            c = conn.cursor()
            c.execute("SELECT * FROM orders WHERE status='pending' AND user_tg_id != 0")
            pending = c.fetchall()
            conn.close()

            for o in pending:
                try:
                    created = datetime.strptime(o['created_at'], "%d.%m.%Y %H:%M")
                    wait_min = int((datetime.now() - created).total_seconds() / 60)
                    if wait_min == 10:
                        try:
                            await bot.send_message(
                                o['user_tg_id'],
                                f"😔 Kechirasiz, #{o['order_uid']} buyurtmangiz hali tasdiqlanmadi. Tez orada hal qilinadi!"
                            )
                        except:
                            pass
                except:
                    pass
        except Exception as e:
            logger.error(f"check_user_order_timeout: {e}")
        await asyncio.sleep(60)

async def check_vacations():
    while True:
        try:
            conn = get_db()
            c = conn.cursor()
            c.execute("SELECT * FROM shops WHERE vacation_until IS NOT NULL AND vacation_until != ''")
            shops = c.fetchall()

            for shop in shops:
                try:
                    vac_date = datetime.strptime(shop['vacation_until'], "%d.%m.%Y").date()
                    if vac_date < date.today():
                        c.execute("UPDATE shops SET vacation_until=NULL, is_open=1 WHERE id=%s", (shop['id'],))
                        try:
                            await bot.send_message(shop['owner_tg_id'],
                                                   f"🏖️ Ta'tilingiz tugadi! {shop['name']} do'koni ishga tushdi!")
                        except:
                            pass
                except:
                    pass

            # Ish vaqtiga qarab avtomatik ochish/yopish
            now_time = datetime.now().strftime("%H:%M")
            c.execute("SELECT id, owner_tg_id, name, work_time, is_open, vacation_until FROM shops WHERE work_time IS NOT NULL AND work_time != ''")
            time_shops = c.fetchall()
            for s in time_shops:
                if s['vacation_until']:
                    continue
                wt = str(s['work_time']).strip().replace(" ", "")
                try:
                    if "-" not in wt:
                        continue
                    open_t, close_t = wt.split("-", 1)
                    should_open = open_t <= now_time < close_t
                    if should_open and not s['is_open']:
                        c.execute("UPDATE shops SET is_open=1 WHERE id=%s", (s['id'],))
                        try:
                            await bot.send_message(s['owner_tg_id'], f"🟢 {s['name']} avtomatik ochildi ({open_t}-{close_t})")
                        except:
                            pass
                    elif not should_open and s['is_open']:
                        c.execute("UPDATE shops SET is_open=0 WHERE id=%s", (s['id'],))
                        try:
                            await bot.send_message(s['owner_tg_id'], f"🔴 {s['name']} avtomatik yopildi ({open_t}-{close_t})")
                        except:
                            pass
                except:
                    pass

            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"check_vacations: {e}")
        await asyncio.sleep(3600)

# --- MAIN ---
async def main():
    init_db()
    asyncio.create_task(check_vacations())
    asyncio.create_task(check_stuck_orders())
    asyncio.create_task(check_user_order_timeout())
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
    
# === ADMIN WEB PANEL ===
from admin_panel import app as web_app
from threading import Thread

def run_web():
    port = int(os.environ.get('PORT', 8080))
    web_app.run(host='0.0.0.0', port=port)

Thread(target=run_web, daemon=True).start()
